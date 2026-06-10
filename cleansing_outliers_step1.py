#!/usr/bin/env python3
"""
清洁工资数据库中的异常日期记录。

用法:
    python cleansing_outliers.py [--dry-run]

参数:
    --dry-run  仅列出异常记录并导出到Excel，不执行删除
    无参数    执行删除操作
"""

import sqlite3
import sys
import argparse
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

DB_PATH = Path(__file__).parent.parent / "payroll_database.db"
OUTPUT_PATH = Path(__file__).parent / "outliers_to_be_deleted.xlsx"
PAYROLL_TABLE = "payroll_details"


def get_outlier_rows(conn: sqlite3.Connection) -> tuple:
    """获取所有异常日期记录。"""
    cursor = conn.execute(
        f"SELECT rowid, * FROM {PAYROLL_TABLE} WHERE 日期 IS NULL OR 日期 = '' OR 日期 GLOB '*：*' OR 日期 GLOB '*月*' OR 日期 LIKE '%加班%' OR 日期 LIKE '%半天%'"
    )
    columns = [description[0] for description in cursor.description]
    rows = cursor.fetchall()
    return columns, rows


def is_row_empty(row_dict: dict, columns: list) -> bool:
    """
    判断该行是否在关键字段（工序全名、工序、计件数量、系数、定额、金额、备注、代码）上都是空值或0。
    如果都是空/0，返回True（real_outliers）；否则返回False（possible_outliers）。
    """
    check_columns = ["工序全名", "工序", "计件数量", "系数", "定额", "金额", "备注", "代码"]

    for col in check_columns:
        if col not in row_dict:
            continue
        value = row_dict[col]

        if value is None or value == "" or value == "None":
            continue

        try:
            if float(value) != 0:
                return False
        except (ValueError, TypeError):
            if str(value).strip() != "":
                return False

    return True


def categorize_outliers(columns: list, rows: list) -> tuple:
    """将异常记录分类为 real_outliers 和 possible_outliers。"""
    real_outliers = []
    possible_outliers = []

    col_indices = {col: idx for idx, col in enumerate(columns)}

    for row in rows:
        row_dict = {col: row[idx] for idx, col in enumerate(columns)}

        if is_row_empty(row_dict, columns):
            real_outliers.append(row)
        else:
            possible_outliers.append(row)

    return real_outliers, possible_outliers


def export_to_excel(columns: list, real_outliers: list, possible_outliers: list, output_path: Path):
    """导出异常记录到Excel文件，包含两个sheet。"""
    wb = openpyxl.Workbook()

    real_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    possible_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def write_sheet(ws, data, sheet_name, row_fill):
        ws.title = sheet_name

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(data, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill

        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["E"].width = 20

    ws1 = wb.active
    write_sheet(ws1, real_outliers, "real_outliers", real_fill)

    ws2 = wb.create_sheet(title="possible_outliers")
    write_sheet(ws2, possible_outliers, possible_outliers and "possible_outliers" or "possible_outliers", possible_fill)

    wb.save(str(output_path))


def delete_outliers(conn: sqlite3.Connection, rows: list) -> int:
    """删除异常记录，返回删除的行数。"""
    if not rows:
        return 0
    rowids = [row[0] for row in rows]
    placeholders = ",".join("?" * len(rowids))
    cursor = conn.execute(
        f"DELETE FROM {PAYROLL_TABLE} WHERE rowid IN ({placeholders})",
        rowids
    )
    conn.commit()
    return cursor.rowcount


def summarize_outliers(columns: list, rows: list) -> dict:
    """生成异常记录统计摘要。"""
    summary = {
        "总数": len(rows),
        "real_outliers": 0,
        "possible_outliers": 0,
        "按文件名分布": {},
        "异常类型分布": {
            "空日期": 0,
            "工序数量误填": 0,
            "月份标注": 0,
            "特殊备注": 0
        }
    }

    date_col_idx = columns.index("日期")
    filename_col_idx = columns.index("文件名")

    real_outliers, possible_outliers = categorize_outliers(columns, rows)
    summary["real_outliers"] = len(real_outliers)
    summary["possible_outliers"] = len(possible_outliers)

    for row in rows:
        date_val = row[date_col_idx] or ""
        filename = row[filename_col_idx]

        summary["按文件名分布"][filename] = summary["按文件名分布"].get(filename, 0) + 1

        if date_val.strip() == "":
            summary["异常类型分布"]["空日期"] += 1
        elif "：" in date_val or ":" in date_val:
            summary["异常类型分布"]["工序数量误填"] += 1
        elif "月" in date_val:
            summary["异常类型分布"]["月份标注"] += 1
        else:
            summary["异常类型分布"]["特殊备注"] += 1

    return summary


def print_summary(summary: dict):
    """打印统计摘要。"""
    print("\n" + "=" * 60)
    print("异常记录统计摘要")
    print("=" * 60)
    print(f"异常记录总数: {summary['总数']}")
    print(f"  - real_outliers (关键字段全为0/空): {summary['real_outliers']}")
    print(f"  - possible_outliers (有有效数据): {summary['possible_outliers']}")

    print("\n按异常类型分布:")
    for outlier_type, count in summary["异常类型分布"].items():
        print(f"  - {outlier_type}: {count}")

    print("\n按文件名分布 (前10个):")
    sorted_files = sorted(summary["按文件名分布"].items(), key=lambda x: x[1], reverse=True)
    for filename, count in sorted_files[:10]:
        print(f"  - {filename}: {count}")

    if len(sorted_files) > 10:
        print(f"  ... 及其他 {len(sorted_files) - 10} 个文件")

    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(
        description="清洁工资数据库中的异常日期记录"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="仅列出异常记录并导出到Excel，不执行删除"
    )
    args = parser.parse_args()

    if not DB_PATH.exists():
        print(f"错误: 数据库文件不存在: {DB_PATH}")
        sys.exit(1)

    conn = sqlite3.connect(str(DB_PATH))

    columns, rows = get_outlier_rows(conn)

    if not rows:
        print("未找到异常记录。")
        conn.close()
        sys.exit(0)

    real_outliers, possible_outliers = categorize_outliers(columns, rows)

    summary = summarize_outliers(columns, rows)
    print_summary(summary)

    export_to_excel(columns, real_outliers, possible_outliers, OUTPUT_PATH)
    print(f"\n异常记录已导出到: {OUTPUT_PATH}")
    print(f"  - real_outliers sheet: {len(real_outliers)} 条")
    print(f"  - possible_outliers sheet: {len(possible_outliers)} 条")

    def count_table_records(conn):
        cursor = conn.execute(f"SELECT COUNT(*) FROM {PAYROLL_TABLE}")
        return cursor.fetchone()[0]

    if args.dry_run:
        print("\n[DRY-RUN 模式] 未执行删除操作。")
    else:
        before_count = count_table_records(conn)
        print(f"\n[确认删除] 数据库当前共有 {before_count} 条记录")
        print(f"[确认删除] 即将删除 {len(real_outliers)} 条 real_outliers 记录...")
        confirm = input("请输入 'yes' 确认删除: ")
        if confirm.strip().lower() == "yes":
            deleted_count = delete_outliers(conn, real_outliers)
            after_count = count_table_records(conn)
            print(f"已成功删除 {deleted_count} 条记录。")
            print(f"删除后数据库共有 {after_count} 条记录。")
        else:
            print("已取消删除操作。")

    conn.close()


if __name__ == "__main__":
    main()
