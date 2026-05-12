#!/usr/bin/env python3
"""
手动检查 Excel 数据问题脚本
逐个打开 Excel 文件，显示错误信息，等待用户确认后继续

使用方法:
    python open_excel_for_review.py
"""

import os
import sys
import subprocess
import shutil
from pathlib import Path

# Base directory
BASE_DIR = Path(__file__).parent.parent.resolve()
TEMP_DIR = BASE_DIR / "temp"
NEW_PAYROLL_DIR = BASE_DIR / "new_payroll"
OLD_PAYROLL_DIR = BASE_DIR / "old_payroll"

# 确保 temp 目录存在
TEMP_DIR.mkdir(parents=True, exist_ok=True)

# List of files with issues from todo.md
# Each file appears only once, all errors from all sheets are shown together
# Format: (file_path, sheet_name, issue_column_letter)
# Sorted by filename ascending
FILES_WITH_ISSUES = [
    # G列（金额列）— #VALUE! 装配喷漆/喷漆装配/精加工
    ("new_payroll/202006.xls", "装配喷漆", "G"),
    ("new_payroll/202007.xls", "装配喷漆", "G"),
    ("new_payroll/202009.xls", "装配喷漆", "G"),
    ("new_payroll/202010.xls", "装配喷漆", "G"),
    ("new_payroll/202011.xls", "装配喷漆", "G"),
    ("new_payroll/202012.xlsx", "喷漆装配", "G"),
    ("new_payroll/202101.xls", "装配喷漆", "G"),
    ("new_payroll/202102.xls", "装配喷漆", "G"),
    ("new_payroll/202105.xls", "装配喷漆", "G"),
    ("new_payroll/202106.xls", "装配喷漆", "G"),
    ("new_payroll/202108.xls", "装配喷漆", "G"),
    ("new_payroll/202110.xls", "精加工", "G"),
    # ("new_payroll/202504.xls", "汇总", "C"),
    # placeholder 目录（旧版备份）
    # ("old_payroll/placeholder/202003.xls", "喷漆装配", "E"),
    # ("old_payroll/placeholder/202006.xls", "装配喷漆", "G"),
    # ("old_payroll/placeholder/202007.xls", "装配喷漆", "G"),
    # ("old_payroll/placeholder/202009.xls", "装配喷漆", "G"),
    # ("old_payroll/placeholder/202010.xls", "装配喷漆", "G"),
    # ("old_payroll/placeholder/202010_2.xls", "装配喷漆", "G"),
    # ("old_payroll/placeholder/202011.xls", "装配喷漆", "G"),
    # 其他
    # ("old_payroll/201711.xls", "装配喷漆", "R"),
]

# 按文件名排序
FILES_WITH_ISSUES.sort(key=lambda x: x[0])


def get_full_path(file_rel_path):
    """根据相对路径获取完整路径"""
    if file_rel_path.startswith('new_payroll'):
        return NEW_PAYROLL_DIR / file_rel_path.replace('new_payroll/', '')
    elif file_rel_path.startswith('old_payroll'):
        return OLD_PAYROLL_DIR / file_rel_path.replace('old_payroll/', '')
    return BASE_DIR / file_rel_path


def convert_xls_to_xlsx(input_path, output_dir=None):
    """使用 LibreOffice 将 .xls 文件转换为 .xlsx"""
    input_path = Path(input_path)
    
    if output_dir is None:
        output_dir = TEMP_DIR
    else:
        output_dir = Path(output_dir)
    
    if input_path.suffix.lower() == '.xlsx':
        return input_path
    
    cmd = [
        'soffice',
        '--headless',
        '--convert-to', 'xlsx',
        '--outdir', str(output_dir),
        str(input_path.resolve())
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode == 0:
            output_path = output_dir / input_path.with_suffix('.xlsx').name
            if output_path.exists():
                return output_path
            else:
                potential_files = list(output_dir.glob(f"{input_path.stem}*.xlsx"))
                if potential_files:
                    return potential_files[0]
    except Exception as e:
        print(f"    转换错误: {e}")
    
    return None


def find_all_errors_in_column(ws, column_letter, start_row=1):
    """查找指定列中所有错误"""
    try:
        from openpyxl.utils import column_index_from_string
    except ImportError:
        return []
    
    col_idx = column_index_from_string(column_letter)
    error_rows = []
    
    for row in range(start_row, min(ws.max_row + 1, start_row + 1000)):
        try:
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell_value = str(cell.value)
                if '#VALUE!' in cell_value or '#REF!' in cell_value:
                    error_rows.append(row)
        except Exception:
            continue
    
    return error_rows


def find_sheet_in_workbook(wb, target_sheet_name):
    """在工作簿中查找工作表"""
    if target_sheet_name in wb.sheetnames:
        return wb[target_sheet_name]
    
    variations = {
        '装配喷漆': ['装配喷漆', '喷漆装配'],
        '喷漆装配': ['装配喷漆', '喷漆装配'],
        '精加工': ['精加工', '金加工'],
        '汇总': ['汇总', '汇总表'],
    }
    
    if target_sheet_name in variations:
        for var in variations[target_sheet_name]:
            if var in wb.sheetnames:
                return wb[var]
    
    for sheet_name in wb.sheetnames:
        if target_sheet_name in sheet_name:
            return wb[sheet_name]
    
    return None


def get_all_sheet_errors(filepath):
    """获取文件中所有工作表的错误信息"""
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter
    
    filepath = Path(filepath)
    error_message = None
    
    if not filepath.exists():
        return False, {}, "文件不存在"
    
    # 如果需要，转换 .xls 为 .xlsx
    if filepath.suffix.lower() == '.xls':
        converted = convert_xls_to_xlsx(filepath)
        if converted is None:
            return False, {}, "转换失败"
        work_file = converted
    else:
        work_file = filepath
    
    try:
        wb = openpyxl.load_workbook(work_file, data_only=True)
        
        all_errors = {}
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 检查所有可能的错误列 (G, L, C, E, F, R)
            error_cols = ['G', 'L', 'C', 'E', 'F', 'R']
            
            for error_col in error_cols:
                col_idx = column_index_from_string(error_col)
                error_rows = []
                
                for row in range(2, min(ws.max_row + 1, 1000)):  # 从第2行开始（跳过表头）
                    try:
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            cell_value = str(cell.value)
                            if '#VALUE!' in cell_value or '#REF!' in cell_value:
                                error_rows.append(row)
                    except Exception:
                        continue
                
                if error_rows:
                    key = f"{sheet_name} - 列{error_col}"
                    all_errors[key] = error_rows
        
        wb.close()
        return True, all_errors, None
        
    except Exception as e:
        error_message = str(e)[:50]  # 截取错误信息的开头部分
        return False, {}, f"XML解析错误: {error_message}"


def get_file_info(filepath, sheet_name, error_col):
    """获取文件的错误信息，返回 (成功标志, 错误行列表, 工作表名)"""
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter
    
    filepath = Path(filepath)
    
    if not filepath.exists():
        return False, [], sheet_name
    
    # 如果需要，转换 .xls 为 .xlsx
    if filepath.suffix.lower() == '.xls':
        converted = convert_xls_to_xlsx(filepath)
        if converted is None:
            return False, [], sheet_name
        work_file = converted
    else:
        work_file = filepath
    
    try:
        wb = openpyxl.load_workbook(work_file, data_only=True)
        
        ws = find_sheet_in_workbook(wb, sheet_name)
        if ws is None:
            wb.close()
            return False, [], sheet_name
        
        ws_title = ws.title
        error_rows = find_all_errors_in_column(ws, error_col)
        
        wb.close()
        return True, error_rows, ws_title
        
    except Exception as e:
        return False, [], sheet_name


def open_file_in_libreoffice(filepath):
    """使用 LibreOffice 打开文件"""
    filepath = Path(filepath).resolve()
    
    if not filepath.exists():
        print(f"  文件未找到: {filepath}")
        return False
    
    # 使用 LibreOffice 打开文件
    cmd = ['soffice', str(filepath)]
    
    try:
        subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return True
    except Exception as e:
        print(f"  打开文件失败: {e}")
        return False


def main():
    import argparse
    
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='Excel 数据问题手动检查工具')
    parser.add_argument('-N', '--no-interact', action='store_true', 
                        help='仅打印错误信息，不等待用户确认')
    args = parser.parse_args()
    
    no_interact = args.no_interact
    
    print("="*60)
    print("Excel 数据问题手动检查工具")
    print("="*60)
    print()
    
    total_files = len(FILES_WITH_ISSUES)
    
    for idx, (file_rel_path, sheet_name, error_col) in enumerate(FILES_WITH_ISSUES, 1):
        print(f"\n{'='*60}")
        print(f"文件 {idx}/{total_files}")
        print(f"{'='*60}")
        
        # 获取完整路径
        filepath = get_full_path(file_rel_path)
        filename = filepath.name
        
        # 生成中文月份描述（基于文件名如 202006）
        year = filename[:4]
        month = filename[4:6]
        if month.isdigit():
            month_desc = f"{year}年{int(month):02d}月"
        else:
            # 处理如 202010_2 这样的情况
            month_desc = f"{year}年特定月份"
        
        print(f"{filename},  {month_desc}的工资文件")
        
        # 获取文件中所有工作表的所有错误
        success, all_errors, error_msg = get_all_sheet_errors(filepath)
        
        if not success:
            print(f"状态: {error_msg if error_msg else '无法读取文件'}")
            print()
            print("按 Enter 继续到下一个文件...")
            input()
            continue
        
        # 显示所有错误信息
        total_errors = 0
        for sheet_key, error_rows in all_errors.items():
            print(f"  {sheet_key}: {len(error_rows)} 处错误")
            print(f"    错误行号: {error_rows}")
            total_errors += len(error_rows)
        
        print()
        print(f"总计: {total_errors} 处错误")
        
        if no_interact:
            print()
            continue  # 直接到下一个文件
        
        print()
        print("-"*60)
        print("请在 LibreOffice 中检查此文件")
        print("按 Y 确认已检查完毕，程序将打开下一个文件")
        print("按 N 跳过（不打开），继续到下一个文件")
        print("按 Q 退出程序")
        print("-"*60)
        
        # 等待用户确认
        while True:
            user_input = input("请输入 (Y/N/Q): ").strip().upper()
            
            if user_input == 'Y':
                print("正在打开文件...")
                open_file_in_libreoffice(filepath)
                print("继续到下一个文件...")
                break
            elif user_input == 'N':
                print("跳过，不打开文件。继续到下一个文件...")
                break
            elif user_input == 'Q':
                print("退出程序")
                return
            else:
                print("无效输入，请输入 Y、N 或 Q")
        
        print()
        
        print()
    
    print(f"\n{'='*60}")
    print("所有文件已检查完毕！")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()