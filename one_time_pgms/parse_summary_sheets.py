"""
parse_summary_sheets.py — 一次性脚本：解析 Excel"汇总"sheet 入库到 payroll_summary

⚠️ 强约束：本脚本对 DB 的写入**仅限于**新表 `payroll_summary`。
- ✅ CREATE TABLE IF NOT EXISTS payroll_summary
- ✅ INSERT INTO payroll_summary
- ❌ 禁止对 payroll_details / load_log / quota / column_seq 任何写操作
- ❌ 禁止 DROP/ALTER 任何表

执行前自动备份 DB 到 ../payroll_database.db.backup_YYYYMMDD_HHMMSS
通过 SQL 写白名单包装函数 safe_execute() 强制只允许白名单写操作。
通过 PRAGMA query_only 在跑完建表后切到只读模式。

支持 --dry-run：仅生成 HTML 预览，不写 DB。
"""
import argparse
import json
import os
import re
import shutil
import sqlite3
import sys
import time
from collections import OrderedDict
from decimal import Decimal
from pathlib import Path

import xlrd
import openpyxl

# 路径
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = PROJECT_ROOT.parent / "payroll_database.db"
OLD_DIR = PROJECT_ROOT / "old_payroll"
NEW_DIR = PROJECT_ROOT / "new_payroll"

# 汇总表 sheet 名变体
SUMMARY_SHEET_NAMES = {"汇总", "统计", "Sheet1"}

# 列名规范化映射（保留原值 + 标准化）
COLUMN_NAME_MAPPING = {
    "车间": "车间",
    "工序": "工序",
    "总件数": "总件数",
}


def find_excel_files():
    """扫描 old_payroll + new_payroll 全部 .xls/.xlsx"""
    files = []
    for d in [OLD_DIR, NEW_DIR]:
        if not d.exists():
            continue
        for p in sorted(d.iterdir()):
            if p.suffix.lower() in (".xls", ".xlsx"):
                files.append(p)
    return files


def get_summary_sheet_names(xls_path):
    """返回该文件中所有疑似"汇总"sheet 的 sheet 名列表"""
    try:
        if str(xls_path).endswith(".xlsx"):
            wb = openpyxl.load_workbook(xls_path, data_only=True, read_only=True)
            names = list(wb.sheetnames)
            wb.close()
        else:
            wb = xlrd.open_workbook(str(xls_path))
            names = wb.sheet_names()
            wb = None
        return [n for n in names if n in SUMMARY_SHEET_NAMES or "汇总" in n]
    except Exception as e:
        return []


def read_sheet_rows(xls_path, sheet_name):
    """读 sheet 所有行为 list[list]，None/空字符串保留"""
    if str(xls_path).endswith(".xlsx"):
        wb = openpyxl.load_workbook(xls_path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
        return rows
    else:
        wb = xlrd.open_workbook(str(xls_path))
        ws = wb.sheet_by_name(sheet_name)
        rows = [ws.row_values(i) for i in range(ws.nrows)]
        return rows


def normalize_row(raw_row, file_name, sheet_name, row_idx, header_row=None):
    """从汇总表一行中抽取 车间/工序/总件数/姓名/工作日/事假/上月件数/累计件数 + 原始列 JSON

    header_row: 第 1 行表头 list（None 时按位置兜底）
    """
    workshop = None
    process = None
    total_qty = None
    prev_qty = None
    cum_qty = None
    name = None
    workdays = None
    leave_days = None
    raw_dict = OrderedDict()

    for i, v in enumerate(raw_row):
        if v is None or (isinstance(v, str) and v.strip() == ""):
            raw_dict[f"col_{i}"] = None
            continue
        sv = str(v).strip() if isinstance(v, str) else v
        raw_dict[f"col_{i}"] = sv

    # 优先按表头列名识别
    if header_row:
        for col_idx, header in enumerate(header_row):
            if header is None:
                continue
            h = str(header).strip()
            if col_idx >= len(raw_row):
                continue
            val = raw_row[col_idx]
            if h == "车间" and val is not None and str(val).strip():
                workshop = str(val).strip()
            elif h == "工序" and val is not None and str(val).strip():
                process = str(val).strip()
            elif h == "总件数" and val is not None:
                try:
                    total_qty = float(val)
                except (ValueError, TypeError):
                    pass
            elif h in ("姓名", "员工", "职员") and val is not None and str(val).strip():
                name = str(val).strip()
            elif "工作日" in h and val is not None:
                try:
                    workdays = float(val)
                except (ValueError, TypeError):
                    pass
            elif "事假" in h and val is not None:
                try:
                    leave_days = float(val)
                except (ValueError, TypeError):
                    pass
            elif ("上月" in h or "5月" in h or "6月" in h or "7月" in h) and "件数" in h and val is not None:
                try:
                    prev_qty = float(val)
                except (ValueError, TypeError):
                    pass
            elif "累计" in h and "件数" in h and val is not None:
                try:
                    cum_qty = float(val)
                except (ValueError, TypeError):
                    pass
    else:
        # 兜底：按位置
        if len(raw_row) >= 1 and raw_row[0] is not None and str(raw_row[0]).strip():
            workshop = str(raw_row[0]).strip()
        if len(raw_row) >= 2 and raw_row[1] is not None and str(raw_row[1]).strip():
            process = str(raw_row[1]).strip()
        if len(raw_row) >= 3 and raw_row[2] is not None:
            try:
                total_qty = float(raw_row[2])
            except (ValueError, TypeError):
                pass

    return {
        "文件名": file_name,
        "sheet名": sheet_name,
        "汇总行索引": row_idx,
        "车间": workshop,
        "工序": process,
        "总件数": total_qty,
        "姓名": name,
        "工作日": workdays,
        "事假": leave_days,
        "上月件数": prev_qty,
        "累计件数": cum_qty,
        "原始列": json.dumps(raw_dict, ensure_ascii=False, default=str),
    }


def safe_execute(conn, sql, params=None, allow_write=False):
    """SQL 白名单包装：只允许 SELECT 和白名单写操作"""
    sql_stripped = " ".join(sql.strip().split()).upper()  # 合并空白 + 大写
    if sql_stripped.startswith("SELECT") or sql_stripped.startswith("PRAGMA"):
        cur = conn.execute(sql, params or ())
        return cur
    if not allow_write:
        raise PermissionError(f"❌ 拒绝非白名单写操作: {sql_stripped[:80]}")
    # 白名单写：CREATE TABLE payroll_summary / INSERT INTO payroll_summary
    allowed_patterns = [
        r"^CREATE\s+TABLE\s+IF\s+NOT\s+EXISTS\s+PAYROLL_SUMMARY",
        r"^INSERT\s+INTO\s+PAYROLL_SUMMARY",
    ]
    if not any(re.match(p, sql_stripped) for p in allowed_patterns):
        raise PermissionError(f"❌ 拒绝非白名单写操作: {sql_stripped[:80]}")
    cur = conn.execute(sql, params or ())
    return cur


def backup_db():
    """备份 DB 到 .backup_YYYYMMDD_HHMMSS"""
    if not DB_PATH.exists():
        print(f"⚠️ DB 不存在: {DB_PATH}")
        return None
    ts = time.strftime("%Y%m%d_%H%M%S")
    backup_path = DB_PATH.parent / f"payroll_database.db.backup_{ts}"
    shutil.copy2(DB_PATH, backup_path)
    print(f"✅ DB 已备份: {backup_path}")
    return backup_path


def create_payroll_summary_table(conn):
    """建表 payroll_summary"""
    sql = """
    CREATE TABLE IF NOT EXISTS payroll_summary (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        文件名 TEXT NOT NULL,
        sheet名 TEXT NOT NULL,
        汇总行索引 INTEGER,
        车间 TEXT,
        工序 TEXT,
        总件数 REAL,
        姓名 TEXT,
        工作日 REAL,
        事假 REAL,
        上月件数 REAL,
        累计件数 REAL,
        原始列 TEXT
    )
    """
    safe_execute(conn, sql, allow_write=True)
    conn.commit()
    print("✅ payroll_summary 表已确保存在")


def insert_payroll_summary_row(conn, row):
    """插入一行"""
    sql = """
    INSERT INTO payroll_summary (文件名, sheet名, 汇总行索引, 车间, 工序, 总件数, 姓名, 工作日, 事假, 上月件数, 累计件数, 原始列)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    safe_execute(
        conn,
        sql,
        (
            row["文件名"],
            row["sheet名"],
            row["汇总行索引"],
            row["车间"],
            row["工序"],
            row["总件数"],
            row["姓名"],
            row["工作日"],
            row["事假"],
            row["上月件数"],
            row["累计件数"],
            row["原始列"],
        ),
        allow_write=True,
    )


def parse_all(dry_run=False):
    """主流程：扫描所有 Excel，找汇总表，规范化入库"""
    files = find_excel_files()
    print(f"📁 扫描到 {len(files)} 个 Excel 文件")

    summary_files = []  # (path, sheet_name)
    for p in files:
        names = get_summary_sheet_names(p)
        for sn in names:
            summary_files.append((p, sn))
    print(f"📋 其中 {len(summary_files)} 个文件含疑似汇总 sheet")

    parsed_rows = []
    skipped_files = []

    for p, sn in summary_files:
        try:
            rows = read_sheet_rows(p, sn)
            file_name = p.name
            # 第 0 行 = 标题，第 1 行 = 表头，从第 2 行起是数据
            header_row = rows[1] if len(rows) > 1 else None
            for i, raw_row in enumerate(rows):
                if i < 2:  # 跳过标题 + 表头
                    continue
                # 过滤全空行
                if all(v is None or (isinstance(v, str) and v.strip() == "") for v in raw_row):
                    continue
                normalized = normalize_row(raw_row, file_name, sn, i, header_row)
                # 必须有车间或工序之一才保留
                if normalized["车间"] or normalized["工序"]:
                    parsed_rows.append(normalized)
        except Exception as e:
            skipped_files.append((p, sn, str(e)))
            print(f"⚠️ 跳过 {p.name}::{sn} - {e}")

    print(f"✅ 解析得到 {len(parsed_rows)} 行汇总数据")
    if skipped_files:
        print(f"⚠️ 跳过 {len(skipped_files)} 个汇总 sheet（读取失败）")

    if dry_run:
        return parsed_rows, skipped_files
    else:
        # 真实入库
        backup_db()
        conn = sqlite3.connect(str(DB_PATH))
        try:
            create_payroll_summary_table(conn)
            for row in parsed_rows:
                insert_payroll_summary_row(conn, row)
            conn.commit()
            # 切到只读模式防误操作
            conn.execute("PRAGMA query_only = ON")
            print(f"✅ 已写入 {len(parsed_rows)} 行到 payroll_summary")
            # 统计
            cur = safe_execute(conn, "SELECT COUNT(*), COUNT(DISTINCT 文件名) FROM payroll_summary")
            total, files_count = cur.fetchone()
            print(f"📊 payroll_summary 总行数: {total} (来自 {files_count} 个文件)")
        finally:
            conn.close()

    return parsed_rows, skipped_files


def generate_html_preview(parsed_rows, skipped_files, output_path):
    """生成 dry-run 预览 HTML"""
    # 按文件汇总
    file_groups = OrderedDict()
    for r in parsed_rows:
        k = r["文件名"]
        if k not in file_groups:
            file_groups[k] = []
        file_groups[k].append(r)

    html = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'>",
        "<title>汇总表解析预览 - parse_summary_sheets</title>",
        "<style>",
        "body{font-family:sans-serif;margin:20px;}",
        "h1{color:#333;}",
        "h2{background:#eef;padding:8px;border-left:4px solid #06c;}",
        "table{border-collapse:collapse;margin:10px 0;font-size:13px;}",
        "th,td{border:1px solid #ccc;padding:4px 8px;text-align:left;}",
        "th{background:#f5f5f5;}",
        ".empty{color:#999;}",
        "</style></head><body>",
        f"<h1>汇总表解析预览 (Dry-Run)</h1>",
        f"<p>共解析 <b>{len(parsed_rows)}</b> 行, 来自 <b>{len(file_groups)}</b> 个文件</p>",
    ]

    if skipped_files:
        html.append("<h2>跳过的文件</h2><ul>")
        for p, sn, err in skipped_files:
            html.append(f"<li>{p.name} :: {sn} - {err}</li>")
        html.append("</ul>")

    for fname, rows in list(file_groups.items())[:5]:
        html.append(f"<h2>{fname}</h2>")
        html.append("<table><tr><th>行</th><th>车间</th><th>工序</th><th>总件数</th><th>姓名</th><th>工作日</th><th>事假</th><th>上月件数</th><th>累计件数</th><th>原始列(前3列)</th></tr>")
        for r in rows[:30]:
            try:
                raw = json.loads(r["原始列"])
                raw_preview = json.dumps({k: v for k, v in list(raw.items())[:3]}, ensure_ascii=False)
            except Exception:
                raw_preview = r["原始列"][:80]
            html.append(
                f"<tr><td>{r['汇总行索引']}</td>"
                f"<td>{r['车间'] or '<span class=empty>空</span>'}</td>"
                f"<td>{r['工序'] or '<span class=empty>空</span>'}</td>"
                f"<td>{r['总件数'] if r['总件数'] is not None else '<span class=empty>空</span>'}</td>"
                f"<td>{r['姓名'] or '<span class=empty>空</span>'}</td>"
                f"<td>{r['工作日'] if r['工作日'] is not None else '<span class=empty>空</span>'}</td>"
                f"<td>{r['事假'] if r['事假'] is not None else '<span class=empty>空</span>'}</td>"
                f"<td>{r['上月件数'] if r['上月件数'] is not None else '<span class=empty>空</span>'}</td>"
                f"<td>{r['累计件数'] if r['累计件数'] is not None else '<span class=empty>空</span>'}</td>"
                f"<td><code>{raw_preview}</code></td></tr>"
            )
        if len(rows) > 30:
            html.append(f"<tr><td colspan=10>... 还有 {len(rows) - 30} 行 ...</td></tr>")
        html.append("</table>")

    html.append("</body></html>")

    Path(output_path).write_text("\n".join(html), encoding="utf-8")
    print(f"✅ 预览已写入: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="解析 Excel 汇总 sheet 入库")
    parser.add_argument("--dry-run", action="store_true", help="仅生成预览，不写 DB")
    parser.add_argument("--output", default="summary_parse_preview.html", help="预览 HTML 路径")
    args = parser.parse_args()

    if not DB_PATH.exists():
        print(f"❌ DB 不存在: {DB_PATH}")
        sys.exit(1)

    parsed_rows, skipped_files = parse_all(dry_run=args.dry_run)
    if args.dry_run:
        generate_html_preview(parsed_rows, skipped_files, args.output)


if __name__ == "__main__":
    main()
