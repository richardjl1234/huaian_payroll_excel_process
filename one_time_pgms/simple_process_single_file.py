import os
import openpyxl
import xlrd
from datetime import datetime

# 简单的单文件处理器，将结果保存到文本文件
def process_single_excel_file(file_path):
    """
    处理单个Excel文件并将详细结果保存到文本文件
    """
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return False
    
    # 创建输出文件名
    output_file = f"process_result_{os.path.basename(file_path).split('.')[0]}.txt"
    
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"=== 开始处理文件: {os.path.basename(file_path)} ===\n")
            f.write(f"文件路径: {file_path}\n")
            f.write(f"处理时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            # 根据文件扩展名选择处理方法
            if file_path.lower().endswith('.xlsx'):
                process_xlsx(file_path, f)
            elif file_path.lower().endswith('.xls'):
                process_xls(file_path, f)
            else:
                f.write(f"不支持的文件格式: {file_path}\n")
                return False
            
            f.write(f"\n=== 处理完成 ===\n")
            f.write(f"结果已保存到: {output_file}\n")
        
        print(f"处理完成，结果已保存到: {output_file}")
        return True
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        return False

def process_xlsx(file_path, output_file):
    """
    处理.xlsx格式的Excel文件
    """
    try:
        # 打开文件
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        output_file.write(f"文件类型: .xlsx\n")
        output_file.write(f"工作表总数: {len(workbook.sheetnames)}\n\n")
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 检查是否为'汇总'表
            if sheet_name == '汇总':
                output_file.write(f"忽略工作表: '汇总'\n")
                continue
            
            # 检查是否为空表
            if sheet.max_row <= 1 and sheet.max_column <= 1:
                output_file.write(f"忽略空工作表: '{sheet_name}'\n")
                continue
            
            output_file.write(f"\n处理工作表: {sheet_name}\n")
            output_file.write(f"工作表尺寸: {sheet.max_row}行 x {sheet.max_column}列\n")
            
            # 尝试提取前5行数据作为示例
            output_file.write("前5行数据示例:\n")
            for i, row in enumerate(sheet.iter_rows(max_row=5), 1):
                cells = []
                for cell in row[:5]:  # 只显示前5列
                    cell_value = "空" if cell.value is None else str(cell.value).strip()
                    cells.append(cell_value)
                output_file.write(f"行{i}: {', '.join(cells)}\n")
    except Exception as e:
        output_file.write(f"处理.xlsx文件时出错: {str(e)}\n")

def process_xls(file_path, output_file):
    """
    处理.xls格式的Excel文件
    """
    try:
        # 打开文件
        workbook = xlrd.open_workbook(file_path)
        output_file.write(f"文件类型: .xls\n")
        output_file.write(f"工作表总数: {workbook.nsheets}\n\n")
        
        # 遍历所有工作表
        for sheet_idx in range(workbook.nsheets):
            sheet_name = workbook.sheet_names()[sheet_idx]
            sheet = workbook.sheet_by_index(sheet_idx)
            
            # 检查是否为'汇总'表
            if sheet_name == '汇总':
                output_file.write(f"忽略工作表: '汇总'\n")
                continue
            
            # 检查是否为空表
            if sheet.nrows <= 1 and sheet.ncols <= 1:
                output_file.write(f"忽略空工作表: '{sheet_name}'\n")
                continue
            
            output_file.write(f"\n处理工作表: {sheet_name}\n")
            output_file.write(f"工作表尺寸: {sheet.nrows}行 x {sheet.ncols}列\n")
            
            # 尝试提取前5行数据作为示例
            output_file.write("前5行数据示例:\n")
            max_rows = min(5, sheet.nrows)
            for row_idx in range(max_rows):
                cells = []
                max_cols = min(5, sheet.ncols)
                for col_idx in range(max_cols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    cell_value = "空" if cell_value == '' or cell_value is None else str(cell_value).strip()
                    cells.append(cell_value)
                output_file.write(f"行{row_idx+1}: {', '.join(cells)}\n")
    except Exception as e:
        output_file.write(f"处理.xls文件时出错: {str(e)}\n")

if __name__ == '__main__':
    # 获取当前工作目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 处理new_payroll中的202001文件
    new_payroll_file = os.path.join(current_dir, 'new_payroll', '202001.xls')
    if os.path.exists(new_payroll_file):
        process_single_excel_file(new_payroll_file)
    else:
        print(f"new_payroll中的202001.xls文件不存在: {new_payroll_file}")
    
    # 处理old_payroll中的202001文件（如果存在）
    old_payroll_file = os.path.join(current_dir, 'old_payroll', '202001.xls')
    if os.path.exists(old_payroll_file):
        process_single_excel_file(old_payroll_file)
    else:
        print(f"old_payroll中的202001.xls文件不存在: {old_payroll_file}")