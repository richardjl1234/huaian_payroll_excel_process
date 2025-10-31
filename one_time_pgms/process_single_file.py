import os
import logging
from datetime import datetime
import openpyxl
import xlrd
import warnings

# 忽略一些警告
warnings.filterwarnings('ignore')

# 设置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()  # 只输出到控制台
    ]
)
logger = logging.getLogger(__name__)

class SingleFileExcelProcessor:
    """
    单文件Excel处理器，用于专门处理202001文件并输出详细结果
    """
    def __init__(self):
        # 设置空行阈值，连续超过此行数视为表格分隔
        self.empty_row_threshold = 2
    
    def is_valid_sheet(self, sheet, sheet_name):
        """
        检查openpyxl工作表是否有效（非空且不是'汇总'表）
        """
        # 忽略名为'汇总'的工作表
        if sheet_name == '汇总':
            logger.warning(f"忽略工作表 '汇总'")
            return False
        
        # 检查工作表是否为空
        if sheet.max_row <= 1 and sheet.max_column <= 1:
            logger.warning(f"工作表 '{sheet_name}' 为空，忽略")
            return False
        
        return True
    
    def is_valid_xls_sheet(self, sheet, sheet_name):
        """
        检查xlrd工作表是否有效（非空且不是'汇总'表）
        """
        # 忽略名为'汇总'的工作表
        if sheet_name == '汇总':
            logger.warning(f"忽略工作表 '汇总'")
            return False
        
        # 检查工作表是否为空
        if sheet.nrows <= 1 and sheet.ncols <= 1:
            logger.warning(f"工作表 '{sheet_name}' 为空，忽略")
            return False
        
        return True
    
    def is_header_row_xlsx(self, row, threshold=3):
        """
        判断openpyxl工作表的一行是否为表格标题行
        """
        non_empty_cells = 0
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != '':
                non_empty_cells += 1
                # 检查是否有粗体文本，标题行通常有粗体
                if hasattr(cell, 'font') and cell.font and cell.font.bold:
                    return True
        
        # 如果非空单元格数量超过阈值，也视为标题行
        return non_empty_cells >= threshold
    
    def process_xlsx_sheet(self, sheet, sheet_name):
        """
        处理openpyxl工作表，提取表格标题行
        """
        headers_set = set()
        current_empty_rows = 0
        in_table = False
        table_count = 0
        
        logger.info(f"\n处理工作表: {sheet_name}")
        logger.info(f"工作表尺寸: {sheet.max_row}行 x {sheet.max_column}列")
        
        # 遍历工作表的每一行
        for row_idx, row in enumerate(sheet.iter_rows(), 1):  # 行索引从1开始
            # 检查是否为空行
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                current_empty_rows += 1
                # 如果连续空行数超过阈值，视为表格结束
                if current_empty_rows >= self.empty_row_threshold and in_table:
                    logger.info(f"  表格{table_count}结束于行{row_idx}")
                    in_table = False
                continue
            else:
                current_empty_rows = 0
            
            # 检查是否为标题行
            if self.is_header_row_xlsx(row):
                if not in_table:
                    table_count += 1
                    logger.info(f"  发现表格{table_count}起始于行{row_idx}")
                in_table = True
                # 提取标题行的列名
                headers = []
                for cell in row:
                    if cell.value is not None:
                        header_text = str(cell.value).strip()
                        if header_text:
                            headers.append(header_text)
                
                # 将标题行转换为用冒号分隔的字符串
                if headers:
                    headers_str = ':'.join(headers)
                    headers_set.add(headers_str)
                    logger.info(f"    表格{table_count}标题行: {headers_str}")
            elif in_table:
                # 表格内容行，继续保持in_table状态
                continue
        
        return headers_set, table_count
    
    def is_header_row_xls(self, sheet, row_idx, threshold=3):
        """
        判断xlrd工作表的一行是否为表格标题行
        """
        non_empty_cells = 0
        
        # xlrd不直接支持字体样式检查，所以我们主要依靠非空单元格数量和内容特征
        for col_idx in range(sheet.ncols):
            cell_value = sheet.cell_value(row_idx, col_idx)
            if cell_value is not None and str(cell_value).strip() != '':
                non_empty_cells += 1
                # 标题通常包含文本，而不是纯数字
                if isinstance(cell_value, str) and len(cell_value.strip()) > 1:
                    # 简单的启发式：如果有多个文本单元格，可能是标题
                    pass
        
        # 如果非空单元格数量超过阈值，视为标题行
        return non_empty_cells >= threshold
    
    def process_xls_sheet(self, sheet, sheet_name):
        """
        处理xlrd工作表，提取表格标题行
        """
        headers_set = set()
        current_empty_rows = 0
        in_table = False
        table_count = 0
        
        logger.info(f"\n处理工作表: {sheet_name}")
        logger.info(f"工作表尺寸: {sheet.nrows}行 x {sheet.ncols}列")
        
        # 遍历工作表的每一行
        for row_idx in range(sheet.nrows):
            # 检查是否为空行
            is_empty = True
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                if cell_value is not None and str(cell_value).strip() != '':
                    is_empty = False
                    break
            
            if is_empty:
                current_empty_rows += 1
                # 如果连续空行数超过阈值，视为表格结束
                if current_empty_rows >= self.empty_row_threshold and in_table:
                    logger.info(f"  表格{table_count}结束于行{row_idx+1}")
                    in_table = False
                continue
            else:
                current_empty_rows = 0
            
            # 检查是否为标题行
            if self.is_header_row_xls(sheet, row_idx):
                if not in_table:
                    table_count += 1
                    logger.info(f"  发现表格{table_count}起始于行{row_idx+1}")
                in_table = True
                # 提取标题行的列名
                headers = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value is not None:
                        header_text = str(cell_value).strip()
                        if header_text:
                            headers.append(header_text)
                
                # 将标题行转换为用冒号分隔的字符串
                if headers:
                    headers_str = ':'.join(headers)
                    headers_set.add(headers_str)
                    logger.info(f"    表格{table_count}标题行: {headers_str}")
            elif in_table:
                # 表格内容行，继续保持in_table状态
                continue
        
        return headers_set, table_count
    
    def process_xlsx_file(self, file_path):
        """
        处理.xlsx格式的Excel文件
        """
        try:
            # 使用openpyxl打开文件
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            valid_sheet_names = []
            all_headers_set = set()
            total_tables = 0
            
            logger.info(f"文件包含 {len(workbook.sheetnames)} 个工作表")
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                if self.is_valid_sheet(sheet, sheet_name):
                    valid_sheet_names.append(sheet_name)
                    # 使用xlsx专用的处理方法
                    headers_set, table_count = self.process_xlsx_sheet(sheet, sheet_name)
                    all_headers_set.update(headers_set)
                    total_tables += table_count
            
            return valid_sheet_names, all_headers_set, total_tables
        except Exception as e:
            logger.error(f"处理.xlsx文件时出错: {str(e)}")
            return [], set(), 0
    
    def process_xls_file(self, file_path):
        """
        处理.xls格式的Excel文件
        """
        try:
            # 使用xlrd打开文件
            workbook = xlrd.open_workbook(file_path)
            valid_sheet_names = []
            all_headers_set = set()
            total_tables = 0
            
            logger.info(f"文件包含 {workbook.nsheets} 个工作表")
            
            for sheet_idx in range(workbook.nsheets):
                sheet_name = workbook.sheet_names()[sheet_idx]
                sheet = workbook.sheet_by_index(sheet_idx)
                
                if self.is_valid_xls_sheet(sheet, sheet_name):
                    valid_sheet_names.append(sheet_name)
                    # 使用xls专用的处理方法
                    headers_set, table_count = self.process_xls_sheet(sheet, sheet_name)
                    all_headers_set.update(headers_set)
                    total_tables += table_count
            
            return valid_sheet_names, all_headers_set, total_tables
        except Exception as e:
            logger.error(f"处理.xls文件时出错: {str(e)}")
            return [], set(), 0
    
    def process_file(self, file_path):
        """
        处理单个Excel文件，根据文件类型选择不同的处理方法
        """
        try:
            # 获取文件名
            file_name = os.path.basename(file_path)
            logger.info(f"\n=== 开始处理文件: {file_name} ===")
            logger.info(f"文件路径: {file_path}")
            
            # 根据文件扩展名选择不同的处理方法
            if file_path.lower().endswith('.xlsx'):
                valid_sheet_names, all_headers_set, total_tables = self.process_xlsx_file(file_path)
            elif file_path.lower().endswith('.xls'):
                valid_sheet_names, all_headers_set, total_tables = self.process_xls_file(file_path)
            else:
                logger.warning(f"不支持的文件格式: {file_path}")
                return
            
            # 打印详细结果
            logger.info(f"\n=== 文件 {file_name} 处理结果 ===")
            logger.info(f"有效工作表数: {len(valid_sheet_names)}")
            logger.info(f"有效工作表名: {', '.join(valid_sheet_names)}")
            logger.info(f"总表格数: {total_tables}")
            logger.info(f"唯一标题数: {len(all_headers_set)}")
            
            # 打印前5个标题示例
            if all_headers_set:
                logger.info(f"标题示例 (前5个):")
                for i, header in enumerate(list(all_headers_set)[:5], 1):
                    logger.info(f"  {i}. {header}")
                
        except Exception as e:
            logger.error(f"处理文件 {file_path} 时出错: {str(e)}")

# 主函数
if __name__ == '__main__':
    start_time = datetime.now()
    
    processor = SingleFileExcelProcessor()
    
    # 获取当前工作目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 处理new_payroll中的202001文件
    new_payroll_file = os.path.join(current_dir, 'new_payroll', '202001.xls')
    if os.path.exists(new_payroll_file):
        processor.process_file(new_payroll_file)
    else:
        logger.warning(f"new_payroll中的202001.xls文件不存在: {new_payroll_file}")
    
    # 处理old_payroll中的202001文件
    old_payroll_file = os.path.join(current_dir, 'old_payroll', '202001.xls')
    if os.path.exists(old_payroll_file):
        processor.process_file(old_payroll_file)
    else:
        logger.warning(f"old_payroll中的202001.xls文件不存在: {old_payroll_file}")
    
    end_time = datetime.now()
    logger.info(f"\n处理完成，总耗时: {end_time - start_time}")