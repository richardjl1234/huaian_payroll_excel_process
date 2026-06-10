#!/usr/bin/env python3
"""
生成 Excel 数据问题报告。
读取 Excel 文件，找出所有错误（#VALUE! / #REF!），生成 HTML 报告。

使用方法:
    python take_excel_screenshot.py
"""

import os
import sys
import subprocess
import shutil
from pathlib import Path
from datetime import datetime

# Base directory
BASE_DIR = Path(__file__).parent.parent.resolve()
REPORT_DIR = BASE_DIR / "screenshot"
TEMP_DIR = BASE_DIR / "temp"
NEW_PAYROLL_DIR = BASE_DIR / "new_payroll"
OLD_PAYROLL_DIR = BASE_DIR / "old_payroll"

# 确保 temp 目录存在
TEMP_DIR.mkdir(parents=True, exist_ok=True)

# List of files with issues from todo.md
# Format: (file_path, sheet_name, issue_column_letter)
FILES_WITH_ISSUES = [
    # G列（金额列）— #VALUE! 装配喷漆/喷漆装配
    ("new_payroll/202011.xls", "装配喷漆", "G"),
    ("new_payroll/202010.xls", "装配喷漆", "G"),
    ("new_payroll/202009.xls", "装配喷漆", "G"),
    ("new_payroll/202006.xls", "喷漆装配", "G"),
    ("new_payroll/202007.xls", "喷漆装配", "G"),
    ("new_payroll/202012.xlsx", "喷漆装配", "G"),
    ("new_payroll/202101.xls", "喷漆装配", "G"),
    ("new_payroll/202102.xls", "喷漆装配", "G"),
    ("new_payroll/202105.xls", "喷漆装配", "G"),
    ("new_payroll/202108.xls", "装配喷漆", "G"),
    
    # G列（金额列）— #VALUE! 精加工
    ("new_payroll/202006.xls", "精加工", "G"),
    ("new_payroll/202106.xls", "精加工", "G"),
    ("new_payroll/202101.xls", "精加工", "G"),
    ("new_payroll/202108.xls", "精加工", "G"),
    ("new_payroll/202105.xls", "精加工", "G"),
    ("new_payroll/202110.xls", "精加工", "G"),
    
    # # placeholder 目录（旧版备份）
    # ("old_payroll/placeholder/202006.xls", "喷漆装配", "G"),
    # ("old_payroll/placeholder/202006.xls", "精加工", "G"),
    # ("old_payroll/placeholder/202007.xls", "喷漆装配", "G"),
    # ("old_payroll/placeholder/202009.xls", "装配喷漆", "G"),
    # ("old_payroll/placeholder/202010.xls", "装配喷漆", "G"),
    # ("old_payroll/placeholder/202011.xls", "装配喷漆", "G"),
    
    # L列（备注列）— #VALUE!
    ("new_payroll/202108.xls", "装配喷漆", "L"),
    ("new_payroll/202106.xls", "喷漆装配", "L"),
    ("new_payroll/202105.xls", "喷漆装配", "L"),
    
    # 汇总表 C/E/F 列 — #REF!
    ("new_payroll/202504.xls", "汇总", "C"),
    
    # 其他
    ("old_payroll/201711.xls", "装配喷漆", "R"),
    # ("old_payroll/placeholder/202003.xls", "喷漆装配", "E"),
    # ("old_payroll/placeholder/202010_2.xls", "装配喷漆", "G"),
]


def ensure_report_dir():
    """创建报告目录"""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"报告目录: {REPORT_DIR}")


def convert_xls_to_xlsx(input_path, output_dir=None):
    """
    使用 LibreOffice 将 .xls 文件转换为 .xlsx
    
    Args:
        input_path: .xls 文件路径
        output_dir: 输出目录（默认与输入相同）
        
    Returns:
        转换后的 .xlsx 文件路径，失败返回 None
    """
    input_path = Path(input_path)
    
    if output_dir is None:
        output_dir = TEMP_DIR  # 默认输出到 temp 目录
    else:
        output_dir = Path(output_dir)
    
    # 如果已经是 .xlsx，跳过
    if input_path.suffix.lower() == '.xlsx':
        return input_path
    
    # 使用 LibreOffice 转换
    cmd = [
        'soffice',
        '--headless',
        '--convert-to', 'xlsx',
        '--outdir', str(output_dir),
        str(input_path.resolve())
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode == 0:
            output_path = output_dir / input_path.with_suffix('.xlsx').name
            if output_path.exists():
                return output_path
            else:
                potential_files = list(output_dir.glob(f"{input_path.stem}*.xlsx"))
                if potential_files:
                    return potential_files[0]
        else:
            print(f"    转换失败: {result.stderr}")
            
    except subprocess.TimeoutExpired:
        print(f"    转换超时")
    except Exception as e:
        print(f"    转换错误: {e}")
    
    return None


def find_all_errors_in_column(ws, column_letter, start_row=1):
    """
    查找指定列中所有 #VALUE! 或 #REF! 错误的行号
    
    Args:
        ws: openpyxl 工作表
        column_letter: 列字母（如 'G', 'L'）
        start_row: 起始行
        
    Returns:
        错误行号列表
    """
    try:
        from openpyxl.utils import column_index_from_string
    except ImportError:
        print("openpyxl 不可用")
        return []
    
    col_idx = column_index_from_string(column_letter)
    error_rows = []
    
    for row in range(start_row, min(ws.max_row + 1, start_row + 1000)):
        try:
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell_value = str(cell.value)
                if '#VALUE!' in cell_value or '#REF!' in cell_value:
                    error_rows.append(row)
        except Exception:
            continue
    
    return error_rows


def find_sheet_in_workbook(wb, target_sheet_name):
    """
    在工作簿中查找工作表，支持名称变体匹配
    
    Args:
        wb: openpyxl 工作簿
        target_sheet_name: 工作表名称
        
    Returns:
        工作表对象或 None
    """
    # 直接匹配
    if target_sheet_name in wb.sheetnames:
        return wb[target_sheet_name]
    
    # 尝试常见变体
    variations = {
        '装配喷漆': ['装配喷漆', '喷漆装配'],
        '喷漆装配': ['装配喷漆', '喷漆装配'],
        '精加工': ['精加工', '金加工'],
        '汇总': ['汇总', '汇总表'],
    }
    
    if target_sheet_name in variations:
        for var in variations[target_sheet_name]:
            if var in wb.sheetnames:
                return wb[var]
    
    # 尝试包含匹配
    for sheet_name in wb.sheetnames:
        if target_sheet_name in sheet_name:
            return wb[sheet_name]
    
    return None


def generate_html_report(ws, error_rows, error_col, filename, sheet_name):
    """
    生成 HTML 报告，显示所有错误位置
    
    Args:
        ws: openpyxl 工作表
        error_rows: 错误行号列表
        error_col: 错误列字母
        filename: 文件名
        sheet_name: 工作表名称
        
    Returns:
        HTML 字符串
    """
    try:
        from openpyxl.utils import get_column_letter, column_index_from_string
    except ImportError:
        return "<html><body>openpyxl 不可用</body></html>"
    
    col_idx = column_index_from_string(error_col)
    first_error_row = error_rows[0] if error_rows else 1
    
    # 计算显示范围（显示第一个错误附近的20行）
    window_size = 20
    start_row = max(1, first_error_row - window_size)
    end_row = min(ws.max_row, first_error_row + window_size)
    
    # 获取表头
    headers = []
    for col in range(1, min(ws.max_column + 1, 25)):
        headers.append(str(ws.cell(row=1, column=col).value or ''))
    
    # 生成 HTML
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>数据问题报告 - {filename}</title>
    <style>
        body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        h1 {{ color: #333; }}
        h2 {{ color: #555; margin-top: 30px; }}
        .summary {{ background: #fff3cd; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
        .error-list {{ background: #fff; padding: 15px; border-radius: 5px; margin: 20px 0; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
        .error-item {{ padding: 8px 0; border-bottom: 1px solid #eee; }}
        .error-item:last-child {{ border-bottom: none; }}
        .error-item:first-child {{ color: #ff0000; font-weight: bold; background: #fff0f0; padding: 10px; border-radius: 3px; }}
        .marker {{ 
            position: fixed; 
            top: 10px; 
            right: 10px; 
            background: #ff0000; 
            color: white; 
            padding: 10px 20px; 
            border-radius: 5px;
            font-weight: bold;
            z-index: 1000;
        }}
        table {{ border-collapse: collapse; background: white; box-shadow: 0 2px 5px rgba(0,0,0,0.1); margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 6px 8px; text-align: left; font-size: 11px; white-space: nowrap; }}
        th {{ background-color: #4CAF50; color: white; position: sticky; top: 0; }}
        tr:hover {{ background-color: #f1f1f1; }}
        .error-cell {{ background-color: #ffcccc !important; color: #cc0000; font-weight: bold; }}
        .error-row {{ background-color: #fff0f0; }}
        .highlight-row {{ background-color: #ffffcc; }}
        .info-box {{
            background: #e7f3ff;
            padding: 10px;
            border-radius: 5px;
            margin: 10px 0;
        }}
    </style>
</head>
<body>
    <div class="marker">⚠ 错误</div>
    <h1>Excel 数据问题报告</h1>
    
    <div class="summary">
        <h2>文件信息</h2>
        <p><strong>文件名:</strong> {filename}</p>
        <p><strong>工作表:</strong> {sheet_name}</p>
        <p><strong>错误列:</strong> {error_col}</p>
        <p><strong>错误数量:</strong> {len(error_rows)} 处</p>
        <p><strong>错误行号:</strong> {', '.join(map(str, error_rows[:50]))}{'...' if len(error_rows) > 50 else ''}</p>
    </div>
    
    <div class="error-list">
        <h2>所有错误详情</h2>
        <p>共发现 <strong>{len(error_rows)}</strong> 处错误：</p>
"""
    
    # 添加错误列表
    for i, row in enumerate(error_rows[:100]):
        is_first = (i == 0)
        cell_value = ws.cell(row=row, column=col_idx).value
        html += f"""        <div class="error-item{' first' if is_first else ''}">
            第 {row} 行: {cell_value}
        </div>
"""
    
    if len(error_rows) > 100:
        html += f"""        <div class="error-item">
            ... 还有 {len(error_rows) - 100} 处错误未显示
        </div>
"""
    
    html += """    </div>
    
    <h2>数据预览（显示第一处错误附近区域）</h2>
    <div class="info-box">
        <strong>显示范围:</strong> 第 """ + f"{start_row} 至 {end_row} 行 | "
    html += f"""<strong>第一处错误位置:</strong> 第 {first_error_row} 行</div>
    
    <div style="overflow-x: auto; max-height: 70vh; overflow-y: auto;">
    <table>
        <thead>
            <tr>
"""
    
    # 添加表头行
    for i, header in enumerate(headers[:20]):
        try:
            col_letter = get_column_letter(i + 1)
        except Exception:
            col_letter = str(i + 1)
        if col_letter == error_col or i + 1 == col_idx:
            html += f'                <th style="background-color: #ff6666;">{col_letter}: {header}</th>\n'
        else:
            html += f'                <th>{col_letter}: {header}</th>\n'
    
    html += """            </tr>
        </thead>
        <tbody>
"""
    
    # 添加数据行
    for row in range(start_row, end_row + 1):
        is_error_row = (row in error_rows)
        row_class = 'error-row' if is_error_row else ''
        
        html += f'            <tr class="{row_class}">\n'
        
        for col in range(1, min(ws.max_column + 1, 21)):
            try:
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value
                
                col_letter = get_column_letter(col)
                td_class = ''
                
                if col == col_idx and is_error_row:
                    td_class = 'error-cell'
                    if cell_value is None:
                        cell_display = '#VALUE!/REF!'
                    else:
                        cell_display = str(cell_value)
                else:
                    if cell_value is None:
                        cell_display = ''
                    else:
                        cell_display = str(cell_value)
                
                # 转义 HTML 特殊字符
                cell_display = cell_display.replace('&', '&').replace('<', '<').replace('>', '>')
                
                html += f'                <td class="{td_class}">{cell_display}</td>\n'
            except Exception:
                html += f'                <td></td>\n'
        
        html += '            </tr>\n'
    
    html += """        </tbody>
    </table>
    </div>
    <script>
        // 自动滚动到错误行区域
        window.onload = function() {
            var errorRow = document.querySelector('.error-row');
            if (errorRow) {
                errorRow.scrollIntoView({ behavior: 'smooth', block: 'center' });
            }
        };
    </script>
</body>
</html>"""
    
    return html


def process_file(filepath, sheet_name, error_col):
    """
    处理单个文件，生成错误报告
    
    Args:
        filepath: Excel 文件路径
        sheet_name: 包含错误的工作表名称
        error_col: 错误列字母
        
    Returns:
        (成功标志, 错误信息列表)
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter
    
    filepath = Path(filepath)
    filename = filepath.name
    
    if not filepath.exists():
        print(f"  文件未找到: {filepath}")
        return False, []
    
    # 如果需要，转换 .xls 为 .xlsx (输出到 temp 目录)
    if filepath.suffix.lower() == '.xls':
        print(f"  正在转换 .xls 为 .xlsx...")
        converted = convert_xls_to_xlsx(filepath, TEMP_DIR)
        if converted is None:
            print(f"  转换失败")
            return False, []
        print(f"  已转换: {converted}")
        work_file = converted
    else:
        work_file = filepath
    
    errors_found = []
    
    try:
        # 加载工作簿（使用 data_only=True 获取缓存的计算值）
        wb = openpyxl.load_workbook(work_file, data_only=True)
    except Exception as e:
        print(f"  无法加载文件（XML解析错误）: {e}")
        print(f"  跳过此文件")
        return False, []
    
    try:
        # 查找目标工作表
        ws = find_sheet_in_workbook(wb, sheet_name)
        if ws is None:
            print(f"  未找到工作表 '{sheet_name}'")
            print(f"    可用工作表: {wb.sheetnames}")
            wb.close()
            return False, []
        
        print(f"  使用工作表: {ws.title}")
        
        # 查找所有错误行
        error_rows = find_all_errors_in_column(ws, error_col)
        
        if not error_rows:
            print(f"  未发现错误")
            wb.close()
            return True, []
        
        print(f"  发现 {len(error_rows)} 处错误")
        
        # 收集错误信息用于终端输出
        col_idx = column_index_from_string(error_col)
        col_letter = get_column_letter(col_idx)
        for row in error_rows:
            cell_value = ws.cell(row=row, column=col_idx).value
            errors_found.append(f"文件: {filename}, 工作表: {sheet_name}, 第{row}行, 第{col_letter}列, 值: {cell_value}")
        
        wb.close()
        
        # 重新加载工作簿（不使用 data_only 以显示公式）
        try:
            wb = openpyxl.load_workbook(work_file, data_only=False)
            ws = find_sheet_in_workbook(wb, sheet_name)
        except Exception as e2:
            print(f"  重新加载文件时出错: {e2}")
            print(f"  将使用第一次加载的数据生成报告")
            # 重新加载第一次的数据（不带公式）
            wb = openpyxl.load_workbook(work_file, data_only=False)
            ws = find_sheet_in_workbook(wb, sheet_name)
        
        # 生成 HTML 报告
        html_content = generate_html_report(ws, error_rows, error_col, filename, sheet_name)
        
        # 保存 HTML
        html_path = REPORT_DIR / f"{filepath.stem}_data_issue.html"
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"  报告已保存: {html_path}")
        
        wb.close()
        return True, errors_found
        
    except Exception as e:
        print(f"  处理文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return False, []


def process_all_files():
    """处理所有文件并生成报告"""
    ensure_report_dir()
    
    success_count = 0
    fail_count = 0
    skip_count = 0
    
    all_errors = []  # 收集所有错误信息用于最终汇总
    
    for file_rel_path, sheet_name, error_col in FILES_WITH_ISSUES:
        print(f"\n{'='*60}")
        print(f"处理: {file_rel_path}")
        print(f"  工作表: {sheet_name}, 错误列: {error_col}")
        
        # 构建完整路径
        if file_rel_path.startswith('new_payroll'):
            filepath = NEW_PAYROLL_DIR / file_rel_path.replace('new_payroll/', '')
        elif file_rel_path.startswith('old_payroll'):
            filepath = OLD_PAYROLL_DIR / file_rel_path.replace('old_payroll/', '')
        else:
            filepath = BASE_DIR / file_rel_path
        
        if not filepath.exists():
            print(f"  跳过: 文件未找到")
            skip_count += 1
            continue
        
        success, errors = process_file(filepath, sheet_name, error_col)
        if success:
            success_count += 1
            all_errors.extend(errors)
        else:
            fail_count += 1
    
    # 输出汇总信息
    print(f"\n{'='*60}")
    print(f"处理完成")
    print(f"  成功: {success_count}")
    print(f"  失败: {fail_count}")
    print(f"  跳过: {skip_count}")
    print(f"  输出目录: {REPORT_DIR}")
    
    # 输出所有错误的汇总
    if all_errors:
        print(f"\n{'='*60}")
        print(f"错误汇总 (共 {len(all_errors)} 处错误):")
        print(f"{'='*60}")
        for error in all_errors:
            print(f"  - {error}")
    
    return all_errors


if __name__ == '__main__':
    all_errors = process_all_files()