import os
import pandas as pd
import logging
from datetime import datetime
import openpyxl
import xlrd
from collections import defaultdict
import warnings

# 忽略一些pandas的警告
warnings.filterwarnings('ignore')

# 设置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("excel_processing.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ImprovedExcelProcessor:
    """
    改进的Excel文件处理器，用于提取Excel文件中的工作表名称和表格标题
    使用'前一行为空时当前行为标题行'的逻辑来识别标题行
    """
    def __init__(self):
        # 初始化结果字典
        self.results = []
        # 当前处理的文件数
        self.processed_files = 0
        # 总文件数
        self.total_files = 0
    
    def is_valid_sheet(self, sheet, sheet_name):
        """
        检查openpyxl工作表是否有效（非空且不是'汇总'表）
        """
        # 忽略名为'汇总'的工作表
        if sheet_name == '汇总':
            logger.warning(f"忽略工作表 '汇总'")
            return False
        
        # 检查工作表是否为空
        if sheet.max_row <= 1 and sheet.max_column <= 1:
            logger.warning(f"工作表 '{sheet_name}' 为空，忽略")
            return False
        
        return True
    
    def is_valid_xls_sheet(self, sheet, sheet_name):
        """
        检查xlrd工作表是否有效（非空且不是'汇总'表）
        """
        # 忽略名为'汇总'的工作表
        if sheet_name == '汇总':
            logger.warning(f"忽略工作表 '汇总'")
            return False
        
        # 检查工作表是否为空
        if sheet.nrows <= 1 and sheet.ncols <= 1:
            logger.warning(f"工作表 '{sheet_name}' 为空，忽略")
            return False
        
        return True
    
    def process_xlsx_sheet(self, sheet, sheet_name):
        """
        处理openpyxl工作表，提取表格标题行
        使用'前一行为空时当前行为标题行'的逻辑来识别标题行
        """
        headers_set = set()
        previous_row_empty = True  # 初始时，前一行视为空
        
        # 遍历工作表的每一行
        for row in sheet.iter_rows():
            # 检查当前行是否为空
            current_row_empty = all(cell.value is None or str(cell.value).strip() == '' for cell in row)
            
            # 如果前一行为空且当前行不为空，则当前行为标题行
            if previous_row_empty and not current_row_empty:
                # 提取标题行的列名
                headers = []
                for cell in row:
                    if cell.value is not None:
                        header_text = str(cell.value).strip()
                        if header_text:
                            headers.append(header_text)
                
                # 将标题行转换为用换行符分隔的字符串
                if headers:
                    headers_str = '\n'.join(headers)
                    headers_set.add(headers_str)
            
            # 更新前一行是否为空的状态
            previous_row_empty = current_row_empty
        
        return headers_set
    
    def process_xls_sheet(self, sheet, sheet_name):
        """
        处理xlrd工作表，提取表格标题行
        使用'前一行为空时当前行为标题行'的逻辑来识别标题行
        """
        headers_set = set()
        previous_row_empty = True  # 初始时，前一行视为空
        
        # 遍历工作表的每一行
        for row_idx in range(sheet.nrows):
            # 检查当前行是否为空
            current_row_empty = True
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                if cell_value is not None and str(cell_value).strip() != '':
                    current_row_empty = False
                    break
            
            # 如果前一行为空且当前行不为空，则当前行为标题行
            if previous_row_empty and not current_row_empty:
                # 提取标题行的列名
                headers = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value is not None:
                        header_text = str(cell_value).strip()
                        if header_text:
                            headers.append(header_text)
                
                # 将标题行转换为用换行符分隔的字符串
                if headers:
                    headers_str = '\n'.join(headers)
                    headers_set.add(headers_str)
            
            # 更新前一行是否为空的状态
            previous_row_empty = current_row_empty
        
        return headers_set
    
    def process_xlsx_file(self, file_path):
        """
        处理.xlsx格式的Excel文件
        """
        try:
            # 使用openpyxl打开文件
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            valid_sheet_names = []
            all_headers_set = set()
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                if self.is_valid_sheet(sheet, sheet_name):
                    valid_sheet_names.append(sheet_name)
                    # 使用xlsx专用的处理方法
                    headers_set = self.process_xlsx_sheet(sheet, sheet_name)
                    all_headers_set.update(headers_set)
            
            return valid_sheet_names, all_headers_set
        except Exception as e:
            logger.error(f"处理.xlsx文件时出错: {str(e)}")
            return [], set()
    
    def process_xls_file(self, file_path):
        """
        处理.xls格式的Excel文件
        """
        try:
            # 使用xlrd打开文件
            workbook = xlrd.open_workbook(file_path)
            valid_sheet_names = []
            all_headers_set = set()
            
            for sheet_idx in range(workbook.nsheets):
                sheet_name = workbook.sheet_names()[sheet_idx]
                sheet = workbook.sheet_by_index(sheet_idx)
                
                if self.is_valid_xls_sheet(sheet, sheet_name):
                    valid_sheet_names.append(sheet_name)
                    # 使用xls专用的处理方法
                    headers_set = self.process_xls_sheet(sheet, sheet_name)
                    all_headers_set.update(headers_set)
            
            return valid_sheet_names, all_headers_set
        except Exception as e:
            logger.error(f"处理.xls文件时出错: {str(e)}")
            return [], set()
    
    def process_file(self, file_path):
        """
        处理单个Excel文件，根据文件类型选择不同的处理方法
        """
        try:
            # 获取文件名
            file_name = os.path.basename(file_path)
            logger.info(f"开始处理文件: {file_name}")
            
            # 根据文件扩展名选择不同的处理方法
            if file_path.lower().endswith('.xlsx'):
                valid_sheet_names, all_headers_set = self.process_xlsx_file(file_path)
            elif file_path.lower().endswith('.xls'):
                valid_sheet_names, all_headers_set = self.process_xls_file(file_path)
            else:
                logger.warning(f"不支持的文件格式: {file_path}")
                self.processed_files += 1
                return
            
            # 将标题行用换行符连接
            headers_joined = '\n'.join(all_headers_set) if all_headers_set else '无有效标题'
            
            # 将结果添加到列表中
            self.results.append({
                'file_name': file_name,
                'sheet_names': ':'.join(valid_sheet_names),
                'headers': headers_joined
            })
            
            logger.info(f"文件 {file_name} 处理完成: {len(valid_sheet_names)} 个工作表，{len(all_headers_set)} 个唯一标题")
            
        except Exception as e:
            logger.error(f"处理文件 {file_path} 时出错: {str(e)}")
        finally:
            self.processed_files += 1
            # 输出进度
            if self.total_files > 0:
                progress = (self.processed_files / self.total_files) * 100
                logger.info(f"进度: {self.processed_files}/{self.total_files} ({progress:.2f}%)")
    
    def process_folder(self, folder_path):
        """
        处理文件夹中的所有Excel文件
        """
        # 检查文件夹是否存在
        if not os.path.exists(folder_path):
            logger.error(f"文件夹不存在: {folder_path}")
            return
        
        # 获取文件夹中的所有Excel文件
        excel_files = []
        for file in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file)
            # 检查是否为Excel文件
            if os.path.isfile(file_path) and file.lower().endswith(('.xls', '.xlsx')):
                excel_files.append(file_path)
        
        # 更新总文件数
        self.total_files += len(excel_files)
        logger.info(f"在文件夹 {folder_path} 中找到 {len(excel_files)} 个Excel文件")
        
        # 处理每个Excel文件
        for file_path in excel_files:
            self.process_file(file_path)
    
    def run(self):
        """
        运行处理器，处理new_payroll和old_payroll文件夹
        """
        start_time = datetime.now()
        logger.info("开始处理Excel文件")
        
        # 获取当前工作目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 处理new_payroll文件夹
        new_payroll_path = os.path.join(current_dir, 'new_payroll')
        self.process_folder(new_payroll_path)
        
        # 处理old_payroll文件夹
        old_payroll_path = os.path.join(current_dir, 'old_payroll')
        self.process_folder(old_payroll_path)
        
        # 将结果转换为DataFrame
        if self.results:
            df = pd.DataFrame(self.results)
            
            # 保存结果到CSV文件
            output_file = "excel_headers_summary_improved.csv"
            try:
                df.to_csv(output_file, index=False, encoding='utf-8-sig')
                logger.info(f"结果已保存到: {output_file}")
            except PermissionError as e:
                logger.error(f"无法保存结果文件: {str(e)}")
                # 尝试使用备用文件名
                backup_output_file = "excel_headers_backup.csv"
                df.to_csv(backup_output_file, index=False, encoding='utf-8-sig')
                logger.info(f"结果已保存到备用文件: {backup_output_file}")
            
            # 打印结果摘要
            logger.info(f"\n处理摘要:")
            logger.info(f"总处理文件数: {self.processed_files}")
            logger.info(f"有效工作表总数: {sum(len(item['sheet_names'].split(':')) for item in self.results if item['sheet_names'])}")
            logger.info(f"唯一标题总数: {sum(1 for item in self.results if item['headers'] != '无有效标题' and item['headers'] != '')}")
        else:
            logger.warning("未找到任何有效文件或工作表")
        
        end_time = datetime.now()
        logger.info(f"处理完成，耗时: {end_time - start_time}")

# 主函数
if __name__ == '__main__':
    processor = ImprovedExcelProcessor()
    processor.run()