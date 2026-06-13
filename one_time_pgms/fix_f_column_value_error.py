#!/usr/bin/env python3
"""
修复 Excel 文件 F 列（定额）的 #VALUE! 错误。

策略：
  1. 用 LibreOffice 把 .xls 转为 .xlsx
  2. 用 openpyxl (data_only=True) 扫描 G 列（金额）找 #VALUE! / #REF! 行
  3. 对每行，把 F 列（定额）单元格改为 0.8，加黄色背景 + 红色字体
  4. 用 LibreOffice 重新计算 .xlsx（让 G 列 `=F*E` 公式的 cached value 同步更新）
  5. 把 .xls 原文件删除，保留 .xlsx

仅处理 装配喷漆 / 喷漆装配 / 精加工 sheet（G 列 = F*E，金额 = 定额 × 计件数量）。
跳过 L 列（工时扩展列）、R 列、汇总 sheet — 那些不在本次修复范围。

用法:
  python one_time_pgms/fix_f_column_value_error.py --dry-run    # 预览
  python one_time_pgms/fix_f_column_value_error.py             # 实际执行
  python one_time_pgms/fix_f_column_value_error.py --files 202011.xls  # 限定文件
"""

import argparse
import shutil
import subprocess
import sys
from pathlib import Path

BASE_DIR = Path(__file__).parent.parent.resolve()
TEMP_DIR = BASE_DIR / "temp"
NEW_PAYROLL_DIR = BASE_DIR / "new_payroll"
OLD_PAYROLL_DIR = BASE_DIR / "old_payroll"

# Sheets that have F→G (定额→金额) pattern with #VALUE! due to broken VLOOKUP
# 注: 精加工 / 金加工 是同一 sheet 的两种命名 (不同文件用不同名)
TARGET_SHEETS = {"装配喷漆", "喷漆装配", "精加工", "金加工"}

FIXED_VALUE = 0.8
FILL_YELLOW = "FFFFFF00"   # RGB 黄色
FONT_RED = "FFFF0000"      # RGB 红色

# 已知有 F 列 #VALUE! 错误的文件 (来自 todo.md, 扫描日期 2026-05-11)
# 格式: (folder, filename) — 同一文件可能出现多次 (如 202108.xls 有 装配喷漆 + 精加工)
# sheet 名不写死, 由脚本动态扫描 (避免 精加工/金加工 别名问题)
KNOWN_FILES_WITH_ERRORS = [
    # 装配喷漆 / 喷漆装配 表
    ('new_payroll', '202011.xls'),
    ('new_payroll', '202010.xls'),
    ('new_payroll', '202009.xls'),
    ('new_payroll', '202006.xls'),
    ('new_payroll', '202007.xls'),
    ('new_payroll', '202012.xlsx'),
    ('new_payroll', '202101.xls'),
    ('new_payroll', '202102.xls'),
    ('new_payroll', '202105.xls'),
    ('new_payroll', '202108.xls'),
    # 精加工 表 (脚本会同时扫描 精加工 和 金加工 sheet)
    ('new_payroll', '202106.xls'),
    ('new_payroll', '202110.xls'),
]

# 文件损坏无法读取, 直接跳过
KNOWN_CORRUPTED_FILES = {
    ('new_payroll', '202003.xls'),
    ('new_payroll', '202109.xls'),
}


def convert_xls_to_xlsx(xls_path, output_dir):
    """用 LibreOffice 把 .xls 转换为 .xlsx. 返回 .xlsx 路径或 None."""
    output_dir.mkdir(parents=True, exist_ok=True)
    cmd = [
        'soffice', '--headless', '--convert-to', 'xlsx',
        '--outdir', str(output_dir),
        str(xls_path.resolve())
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        print(f"    LibreOffice 转换失败: {result.stderr[:200]}")
        return None
    expected = output_dir / xls_path.with_suffix('.xlsx').name
    if expected.exists():
        return expected
    candidates = list(output_dir.glob(f"{xls_path.stem}*.xlsx"))
    return candidates[0] if candidates else None


def recalc_xlsx(xlsx_path):
    """用 LibreOffice 重新计算 .xlsx 里的所有公式, 更新 cached values.

    必须满足两个条件:
    1) explicit filter "xlsx:Calc Office Open XML" — 隐式 filter 留下空的 <v></v> 标签
    2) outdir 与 xlsx_path.parent 必须不同 — 否则 LibreOffice 写保存失败
       (SfxBaseModel::impl_store failed 0x4c0c, 同路径 in==out)
    """
    outdir = xlsx_path.parent / "_recalc_tmp"
    outdir.mkdir(exist_ok=True)
    try:
        cmd = [
            'soffice', '--headless', '--calc', '--convert-to', 'xlsx:Calc Office Open XML',
            '--outdir', str(outdir),
            str(xlsx_path.resolve())
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode != 0:
            return False
        # 把 outdir 里的 .xlsx 移回原路径
        produced = outdir / xlsx_path.name
        if not produced.exists():
            return False
        shutil.move(str(produced), str(xlsx_path))
        return True
    finally:
        # 清理临时目录
        if outdir.exists():
            shutil.rmtree(outdir, ignore_errors=True)


def find_error_rows(ws_data, error_col='G'):
    """在 G 列(默认) 找值为 '#VALUE!' 或 '#REF!' 的行号列表."""
    from openpyxl.utils import column_index_from_string
    col_idx = column_index_from_string(error_col)
    errors = []
    for row in range(1, ws_data.max_row + 1):
        v = ws_data.cell(row=row, column=col_idx).value
        if v is None:
            continue
        sv = str(v)
        if '#VALUE!' in sv or '#REF!' in sv:
            errors.append(row)
    return errors


def list_target_sheets(xlsx_path):
    """列出 .xlsx 中所有属于 TARGET_SHEETS 的工作表名."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        return None, f"无法读取 ({type(e).__name__})"
    try:
        return [s for s in wb.sheetnames if s in TARGET_SHEETS], None
    finally:
        wb.close()


def process_file_sheets(src, sheets, dry_run=True):
    """
    处理一个源 Excel 文件, 一次性修复所有目标 sheet 的 F 列错误.

    流程:
      .xls  -> LibreOffice 转 .xlsx (temp) -> openpyxl 改 F + 格式 -> save
            -> LibreOffice recalc -> 替换原 .xls
      .xlsx -> openpyxl 改 F + 格式 -> save -> LibreOffice recalc -> 原地保存

    返回 (success, total_modified, error_message, sheet_results)
      sheet_results: list of (sheet_name, modified_count) for reporting
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import column_index_from_string

    if not src.exists():
        return False, 0, "文件不存在", []

    suffix = src.suffix.lower()
    is_xls = (suffix == '.xls')

    # 1) 准备 .xlsx 工作副本
    if is_xls:
        work_xlsx = TEMP_DIR / src.with_suffix('.xlsx').name
        if work_xlsx.exists():
            work_xlsx.unlink()
        if dry_run:
            print(f"    [DRY-RUN] 需 .xls → .xlsx 转换: {src.name}")
        work_xlsx = convert_xls_to_xlsx(src, TEMP_DIR)
        if not work_xlsx:
            return False, 0, ".xls→.xlsx 转换失败", []
        if not dry_run:
            print(f"    转换: {work_xlsx.name}")
    else:
        work_xlsx = src

    # 2) 用 data_only=True 找出所有目标 sheet 的 error rows
    try:
        wb_d = load_workbook(work_xlsx, data_only=True, read_only=True)
    except Exception as e:
        return False, 0, f"读取失败 ({type(e).__name__}: {str(e)[:100]})", []

    sheet_error_rows = {}  # sheet_name -> [error_rows]
    try:
        for sheet in sheets:
            if sheet not in wb_d.sheetnames:
                print(f"    [WARN] 工作表 {sheet} 不存在, 跳过")
                continue
            error_rows = find_error_rows(wb_d[sheet])
            sheet_error_rows[sheet] = error_rows
            if error_rows:
                print(f"    {sheet}: 发现 {len(error_rows)} 行 G 列 #VALUE!/#REF!")
    finally:
        wb_d.close()

    sheet_results = [(s, len(rows)) for s, rows in sheet_error_rows.items() if rows]
    total_modified = sum(c for _, c in sheet_results)

    # dry-run: 直接退出, 不修改文件也不 recalc
    if dry_run:
        return True, total_modified, None, sheet_results

    # 3) 加载并应用修改 (所有 sheet 一次加载, 一次保存)
    wb = load_workbook(work_xlsx, data_only=False)
    f_col = column_index_from_string('F')
    fill = PatternFill(start_color=FILL_YELLOW, end_color=FILL_YELLOW, fill_type='solid')
    font = Font(color=FONT_RED, bold=True)

    # 关键: 关掉 fullCalcOnLoad — 否则 LibreOffice 不会写 G 列公式 cached value
    # (workbook.xml 的 <calcPr fullCalcOnLoad="1"/> 会让重算时故意不写 <v> 标签)
    wb.calculation.fullCalcOnLoad = False

    # 修复 error rows (如果存在)
    for sheet, error_rows in sheet_error_rows.items():
        if not error_rows:
            continue
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in error_rows:
            cell = ws.cell(row=row, column=f_col)
            cell.value = FIXED_VALUE
            cell.fill = fill
            cell.font = font

    wb.save(work_xlsx)
    wb.close()

    # 如果没有 error rows, 也需要 recalc 把 G 列 cached value 写回 (.xlsx 原文件常见问题)
    # 但若 total_modified==0, 提前退出到主流程, 主流程会跳过 recalc — 这里让 recalc 总是跑

    # 4) LibreOffice recalc (让所有 G 列 `=F*E` 公式的 cached value 同步更新)
    print(f"    LibreOffice recalc...")
    if not recalc_xlsx(work_xlsx):
        return False, total_modified, "LibreOffice recalc 失败", sheet_results

    # 5) 把 .xls 替换为 .xlsx
    if is_xls:
        new_path = src.with_suffix('.xlsx')
        shutil.move(str(work_xlsx), str(new_path))
        src.unlink()
        print(f"    已替换: {src.name} → {new_path.name}")

    return True, total_modified, None, sheet_results


def build_work_list(scan_all, only_files=None):
    """
    构建工作清单: list of (src_path, [sheet_names_or_None])
    - scan_all=False: 用 KNOWN_FILES_WITH_ERRORS, sheets 来自列表
    - scan_all=True: 扫描所有 .xls/.xlsx, sheets=None (后续动态扫描)
    - only_files: 只保留文件名匹配的项
    """
    if scan_all:
        paths = []
        for folder in [NEW_PAYROLL_DIR, OLD_PAYROLL_DIR]:
            if folder.exists():
                for f in folder.glob("*.xls*"):
                    if f.is_file():
                        paths.append(f)
        paths.sort()
        items = [(p, None) for p in paths]
    else:
        # 用 todo.md 的已知文件列表
        # KNOWN_FILES_WITH_ERRORS 写的是 .xls 名字, 但跑过一次后 .xls 已替换成 .xlsx
        # 检查两种后缀, 让脚本可以幂等运行
        items = []
        for folder, filename in KNOWN_FILES_WITH_ERRORS:
            base = NEW_PAYROLL_DIR if folder == 'new_payroll' else OLD_PAYROLL_DIR
            xls_path = base / filename
            xlsx_path = xls_path.with_suffix('.xlsx')
            if xls_path.exists():
                items.append((xls_path, None))
            elif xlsx_path.exists():
                items.append((xlsx_path, None))
            else:
                items.append((xls_path, None))  # 让 main 报 "文件不存在"

    if only_files:
        wanted = set(only_files)
        items = [(p, s) for p, s in items if p.name in wanted]
    return items


def main():
    parser = argparse.ArgumentParser(
        description="修复 Excel 文件 F 列 (定额) 的 #VALUE! 错误")
    parser.add_argument('--dry-run', action='store_true',
                        help='预览模式, 不实际修改文件 (但仍会 .xls→.xlsx 转换用于扫描)')
    parser.add_argument('--files', nargs='*', default=None,
                        help='限定只处理指定文件名 (如 202011.xls 202110.xls)')
    parser.add_argument('--all', action='store_true', dest='scan_all',
                        help='扫描所有 .xls/.xlsx, 不只 KNOWN_FILES_WITH_ERRORS 列表里的')
    args = parser.parse_args()

    print(f"模式: {'DRY-RUN' if args.dry_run else 'LIVE'}")
    work_list = build_work_list(args.scan_all, args.files)
    print(f"工作清单: {len(work_list)} 个文件")
    if args.files:
        print(f"  限定文件名: {args.files}")
    if args.scan_all:
        print(f"  扫描范围: 全部 .xls/.xlsx")
    else:
        print(f"  扫描范围: KNOWN_FILES_WITH_ERRORS (来自 todo.md)")
    print()

    total_files_modified = 0
    total_cells_modified = 0
    failed = []
    skipped = 0

    for src, sheets_hint in work_list:
        # 跳过已知损坏文件
        folder_name = 'new_payroll' if 'new_payroll' in str(src) else 'old_payroll'
        if (folder_name, src.name) in KNOWN_CORRUPTED_FILES:
            print(f"[SKIP] {src.relative_to(BASE_DIR)} (已知损坏文件)")
            skipped += 1
            continue

        if not src.exists():
            print(f"[WARN] 文件不存在: {src.relative_to(BASE_DIR)}")
            skipped += 1
            continue

        # 准备临时 .xlsx (如果需要) — 仅用于动态扫描 sheet 名
        if src.suffix.lower() == '.xls':
            tmp = TEMP_DIR / src.with_suffix('.xlsx').name
            if not tmp.exists():
                tmp = convert_xls_to_xlsx(src, TEMP_DIR)
            scan_path = tmp
        else:
            scan_path = src

        if not scan_path or not scan_path.exists():
            print(f"[WARN] 无法准备 {src.name}, 跳过")
            skipped += 1
            continue

        # 决定要处理的 sheet 列表
        if sheets_hint is not None:
            target_sheets = sheets_hint
        else:
            target_sheets, err = list_target_sheets(scan_path)
            if target_sheets is None:
                print(f"[WARN] {src.name}: {err}, 跳过")
                skipped += 1
                continue
            if not target_sheets:
                continue  # 静默跳过没有目标 sheet 的文件

        print(f"=== {src.relative_to(BASE_DIR)} ===")
        for sheet in target_sheets:
            print(f"  -- {sheet} --")
        success, modified, err, sheet_results = process_file_sheets(
            src, target_sheets, dry_run=args.dry_run)
        if not success:
            print(f"    FAILED: {err}")
            failed.append((src.name, err))
        else:
            total_cells_modified += modified
            if modified > 0:
                total_files_modified += 1

    print(f"\n{'='*60}")
    print(f"工作清单: {len(work_list)} 个文件")
    print(f"  跳过 (损坏/无法读取): {skipped}")
    print(f"  失败: {len(failed)}")
    if failed:
        for f, e in failed:
            print(f"    - {f}: {e}")
    print(f"  涉及修复的文件: {total_files_modified}")
    print(f"  修复的 F 列单元格数: {total_cells_modified}")
    print()
    if args.dry_run:
        print("[DRY-RUN 模式] 上面是预览, 未实际修改任何源文件")
    else:
        print("[LIVE 模式] 已就地修改源文件 (.xls 已替换为 .xlsx)")


if __name__ == '__main__':
    main()
