import os
import openpyxl
import xlrd
import pandas as pd
import sqlite3
import logging
from decimal import Decimal, ROUND_HALF_UP
from typing import List
try:
    from .special_logic import special_logic_preprocess_df
    from .config import expected_columns, COMMON_COL_COUNT, setup_global_logging, DATABASE_PATH
except ImportError:
    from special_logic import special_logic_preprocess_df
    from config import expected_columns, COMMON_COL_COUNT, setup_global_logging, DATABASE_PATH

# Set up logging using global configuration
setup_global_logging()
logger = logging.getLogger(__name__)

def _process_cell_value(cell_value):
    """
    Process a cell value to ensure proper string representation.
    For numeric values that should be strings (like model numbers),
    remove the decimal point if it's a whole number.
    
    Parameters:
        cell_value: The raw cell value from Excel
        
    Returns:
        str: Processed string value
    """
    if cell_value is None:
        return ''
    elif isinstance(cell_value, float) and cell_value != cell_value:  # Check for NaN
        return ''
    else:
        # For numeric values that should be strings (like model numbers),
        # remove the decimal point if it's a whole number
        if isinstance(cell_value, (int, float)):
            # Check if it's a whole number
            if cell_value == int(cell_value):
                return str(int(cell_value))
            else:
                return str(cell_value)
        else:
            return str(cell_value)


def get_all_data_from_sheet(excel_file_name, sheet_name):
    """
    Extract all data from an Excel sheet and return as a summary dataframe.
    
    Parameters:
        excel_file_name (str): The name of the Excel file
        sheet_name (str): The name of the sheet to process
        
    Returns:
        pd.DataFrame: Summary dataframe with all contents of the sheet
        
    Note:
        For .xls files, xlrd will evaluate all formulas and return calculated values.
        Formula display values (like #VALUE! errors) cannot be preserved with xlrd.
    """
    # Get the current directory and construct file paths
    current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    new_payroll_path = os.path.join(current_dir, 'new_payroll')
    old_payroll_path = os.path.join(current_dir, 'old_payroll')
    
    # Determine which folder contains the file
    file_path = None
    if os.path.exists(os.path.join(new_payroll_path, excel_file_name)):
        file_path = os.path.join(new_payroll_path, excel_file_name)
    elif os.path.exists(os.path.join(old_payroll_path, excel_file_name)):
        file_path = os.path.join(old_payroll_path, excel_file_name)
    else:
        raise FileNotFoundError(f"Excel file '{excel_file_name}' not found in new_payroll or old_payroll folders")
    
    try:
        # Initialize variables
        all_data = []
        headers = []
        
        # Process based on file type
        if excel_file_name.lower().endswith('.xlsx'):
            # Use openpyxl for .xlsx files
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook[sheet_name]
            
            # Get headers (first row)
            if sheet.max_row > 0:
                for cell in sheet[1]:
                    headers.append(_process_cell_value(cell.value))
            
            # Get all data rows
            for row in sheet.iter_rows(min_row=1, values_only=True):
                row_data = [_process_cell_value(cell_value) for cell_value in row]
                all_data.append(row_data)
                
        elif excel_file_name.lower().endswith('.xls'):
            # Use xlrd for .xls files
            logger.warning(f"Processing .xls file '{excel_file_name}': Formulas will be evaluated and calculated values returned. Formula display values (like #VALUE! errors) cannot be preserved.")
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_name(sheet_name)
            
            # Get headers (first row)
            if sheet.nrows > 0:
                for j in range(sheet.ncols):
                    cell_value = sheet.cell_value(0, j)
                    headers.append(_process_cell_value(cell_value))
            
            # Get all data rows
            for i in range(sheet.nrows):
                row_data = [_process_cell_value(sheet.cell_value(i, j)) for j in range(sheet.ncols)]
                all_data.append(row_data)
        
        # Create df_summary - the complete sheet data
        if all_data:
            # Ensure consistent column lengths
            max_length = max(len(row) for row in all_data)
            extended_headers = headers + [f'Column_{i}' for i in range(len(headers), max_length)]
            
            # Extend all rows to max length
            for i in range(len(all_data)):
                if len(all_data[i]) < max_length:
                    all_data[i].extend([''] * (max_length - len(all_data[i])))
            
            # Handle duplicate column names
            unique_headers = []
            header_counts = {}
            
            for header in extended_headers:
                if header in header_counts:
                    header_counts[header] += 1
                    unique_header = f"{header}_{header_counts[header]}"
                    unique_headers.append(unique_header)
                else:
                    header_counts[header] = 0
                    unique_headers.append(header)
            
            # Create summary DataFrame
            df_summary = pd.DataFrame(all_data, columns=unique_headers)
        else:
            df_summary = pd.DataFrame()
        
        return df_summary
        
    except Exception as e:
        raise Exception(f"Error processing sheet '{sheet_name}' in file '{excel_file_name}': {str(e)}")


def split_raw_sheet_contents(df_summary):
    """
    Split raw sheet contents dataframe into multiple dataframes based on blank rows.
    
    Parameters:
        df_summary (pd.DataFrame): The raw sheet contents as a DataFrame
        
    Returns:
        list: List of pandas DataFrames split by blank rows
    """
    # Get all_data from the dataframe for further processing
    all_data = df_summary.values.tolist()
    unique_headers = df_summary.columns.tolist()
    
    try:
        # Split into multiple dataframes based on blank rows
        dfs = []
        current_data = []
        
        for row in all_data:
            # Check if row is blank (all empty strings)
            if all(cell == '' for cell in row):
                # If we have data collected, create a dataframe
                if current_data:
                    # Ensure consistent column lengths
                    max_len = max(len(r) for r in current_data)
                    for i in range(len(current_data)):
                        if len(current_data[i]) < max_len:
                            current_data[i].extend([''] * (max_len - len(current_data[i])))
                    
                    # Create DataFrame with proper column headers
                    # Use the first row as headers if it contains text data
                    if current_data and any(cell.strip() for cell in current_data[0]):
                        # Use first row as headers
                        headers_for_df = current_data[0]
                        data_for_df = current_data[1:]
                    else:
                        # Use the original headers from the sheet
                        headers_for_df = unique_headers[:max_len]
                        data_for_df = current_data
                    
                    # Handle duplicate column names
                    unique_headers_for_df = []
                    header_counts = {}
                    for header in headers_for_df:
                        if header in header_counts:
                            header_counts[header] += 1
                            unique_header = f"{header}_{header_counts[header]}"
                            unique_headers_for_df.append(unique_header)
                        else:
                            header_counts[header] = 0
                            unique_headers_for_df.append(header)
                    
                    df = pd.DataFrame(data_for_df, columns=unique_headers_for_df)
                    dfs.append(df)
                    current_data = []
            else:
                current_data.append(row)
        
        # Don't forget the last data segment
        if current_data:
            max_len = max(len(r) for r in current_data)
            for i in range(len(current_data)):
                if len(current_data[i]) < max_len:
                    current_data[i].extend([''] * (max_len - len(current_data[i])))
            
            # Create DataFrame with proper column headers
            # Use the first row as headers if it contains text data
            if current_data and any(cell.strip() for cell in current_data[0]):
                # Use first row as headers
                headers_for_df = current_data[0]
                data_for_df = current_data[1:]
            else:
                # Use the original headers from the sheet
                headers_for_df = unique_headers[:max_len]
                data_for_df = current_data
            
            # Handle duplicate column names
            unique_headers_for_df = []
            header_counts = {}
            for header in headers_for_df:
                if header in header_counts:
                    header_counts[header] += 1
                    unique_header = f"{header}_{header_counts[header]}"
                    unique_headers_for_df.append(unique_header)
                else:
                    header_counts[header] = 0
                    unique_headers_for_df.append(header)
            
            df = pd.DataFrame(data_for_df, columns=unique_headers_for_df)
            dfs.append(df)
        
        # Apply column validation logic to each dataframe
        validated_dfs = []
        for df in dfs:
            validated_df = _validate_and_fix_dataframe_columns(df)
            if validated_df is not None:
                validated_dfs.append(validated_df)
        
        return validated_dfs
        
    except Exception as e:
        raise Exception(f"Error splitting raw sheet contents: {str(e)}")


def _validate_and_fix_dataframe_columns(df):
    """
    Validate and fix dataframe columns based on expected_columns and COMMON_COL_COUNT.
    
    Parameters:
        df (pd.DataFrame): The dataframe to validate
        
    Returns:
        pd.DataFrame: The validated and potentially fixed dataframe, or None if invalid
    """
    try:
        # Count how many expected columns are found in the current dataframe
        found_columns = [col for col in df.columns if col in expected_columns]
        found_count = len(found_columns)
        
        # If we have enough expected columns, return the dataframe as is
        if found_count >= COMMON_COL_COUNT:
            return df
        
        # If we don't have enough expected columns, check if any row has more expected columns
        logger.warning(f"Dataframe has only {found_count} expected columns, checking rows for better headers...")
        
        # Iterate through all rows to find a row that contains more expected columns
        best_row_index = -1
        best_row_count = 0
        
        for row_idx, row in df.iterrows():
            # Convert row values to strings and check if they match expected columns
            row_values = [str(val).strip() for val in row.values]
            row_expected_count = sum(1 for val in row_values if val in expected_columns)
            
            if row_expected_count > best_row_count:
                best_row_count = row_expected_count
                best_row_index = row_idx
        
        # If we found a row with more expected columns than current headers
        if best_row_count > found_count and best_row_count >= COMMON_COL_COUNT:
            logger.info(f"Found better headers at row {best_row_index} with {best_row_count} expected columns")
            
            # Use the best row as new headers and all subsequent rows as data
            new_headers = [str(val).strip() for val in df.iloc[best_row_index].values]
            new_data = df.iloc[best_row_index + 1:].reset_index(drop=True)
            
            # Create new dataframe with the new headers
            new_df = pd.DataFrame(new_data.values, columns=new_headers)
            
            # Handle duplicate column names in the new headers
            unique_new_headers = []
            header_counts = {}
            for header in new_headers:
                if header in header_counts:
                    header_counts[header] += 1
                    unique_header = f"{header}_{header_counts[header]}"
                    unique_new_headers.append(unique_header)
                else:
                    header_counts[header] = 0
                    unique_new_headers.append(header)
            
            new_df.columns = unique_new_headers
            return new_df
        else:
            # No suitable row found, log warning and return None to discard this dataframe
            logger.warning(f"No suitable row found with at least {COMMON_COL_COUNT} expected columns. Discarding dataframe.")
            logger.warning(f"Discarded dataframe content:\n{df.to_string()}")
            return None
            
    except Exception as e:
        logger.error(f"Error validating dataframe columns: {str(e)}")
        return df  # Return original dataframe if validation fails



def load_df_to_db(df: pd.DataFrame, file_name: str, sheet_name: str, table_index: int = 0) -> str:
    """
    Load a dataframe to SQLite database with the specified table structure.
    
    Parameters:
        df (pd.DataFrame): The dataframe to load
        file_name (str): The name of the Excel file
        sheet_name (str): The name of the sheet being processed
        table_index (int): The index of the table being processed (e.g., 1 for 表一)
        
    Returns:
        str: Success message or error message
    """
    try:
        # 在函数开始时，移除df.columns中的所有空格和sheet_name中的空格
        # Remove blanks from df.columns
        df.columns = [col.replace(' ', '') for col in df.columns]
        
        # Remove blanks from sheet_name
        clean_sheet_name = sheet_name.replace(' ', '')
        
        # Remove columns that are empty strings or have suffixes like "_1", "_2", etc.
        columns_to_remove = []
        for col in df.columns:
            # Remove empty string columns
            if col == "":
                columns_to_remove.append(col)
            # Remove columns that are just numbers (like "1", "2", etc.)
            elif col.isdigit():
                columns_to_remove.append(col)
            # Remove columns ending with _1, _2, _3, etc.
            elif col.endswith(tuple([f"_{i}" for i in range(1, 100)])):
                columns_to_remove.append(col)
            # Remove columns that start with underscore followed by digits
            elif col.startswith("_") and col[1:].isdigit():
                columns_to_remove.append(col)
        
        # Remove the identified columns
        if columns_to_remove:
            df = df.drop(columns=columns_to_remove)
            logger.info(f"Removed columns: {columns_to_remove}")
        
        # 调用特殊逻辑预处理函数
        # Call special logic preprocessing function
        df, sheet_name, file_name = special_logic_preprocess_df(df, sheet_name, file_name, table_index)
        
        # Define the expected columns and their data types
        expected_columns = {
            '文件名': 'CHAR(100)',
            'sheet名': 'CHAR(100)',
            '职员全名': 'CHAR(20)',
            '日期': 'CHAR(20)', 
            '客户名称': 'CHAR(60)',
            '型号': 'CHAR(100)',
            '工序全名': 'CHAR(100)',
            '工序': 'CHAR(100)',
            '计件数量': 'NUMERIC(10,2)',
            '系数': 'NUMERIC(10,2)',
            '定额': 'NUMERIC(10,2)',
            '金额': 'NUMERIC(10,2)',
            '备注': 'CHAR(100)',
            '代码': 'CHAR(12)'
        }
        
        # Filter dataframe to only include columns that exist in expected columns
        valid_columns = [col for col in df.columns if col in expected_columns.keys()]
        discarded_columns = [col for col in df.columns if col not in expected_columns.keys()]
        discarded_cols_num = len(discarded_columns)
        
        # Check if there are any discarded columns with non-empty names (not '', '_1', '_2', etc.)
        non_empty_discarded_columns = []
        for col in discarded_columns:
            # Check if column name is not empty and doesn't match the pattern of suffixes
            if col != "" and not col.isdigit() and not col.endswith(tuple([f"_{i}" for i in range(1, 100)])) and not (col.startswith("_") and col[1:].isdigit()):
                non_empty_discarded_columns.append(col)
        
        # Log discarded columns to load_log table ONLY if there are discarded columns
        if discarded_columns:
            # Connect to SQLite database using the configured path
            conn = sqlite3.connect(DATABASE_PATH)
            
            # Create load_log table if it doesn't exist
            create_log_table_sql = """
            CREATE TABLE IF NOT EXISTS load_log (
                file_name CHAR(50),
                sheet_name CHAR(50),
                table_index INT,
                discarded_columns CHAR(200),
                discarded_cols_num INT
            )
            """
            conn.execute(create_log_table_sql)
            
            # Insert log record
            discarded_columns_str = ', '.join(discarded_columns)
            insert_log_sql = """
            INSERT INTO load_log (file_name, sheet_name, table_index, discarded_columns, discarded_cols_num)
            VALUES (?, ?, ?, ?, ?)
            """
            conn.execute(insert_log_sql, (file_name, sheet_name, table_index, discarded_columns_str, discarded_cols_num))
            conn.commit()
            conn.close()
            
            # Log warning message about discarded columns
            logger.warning(f"{discarded_cols_num} columns discarded in file '{file_name}', sheet '{sheet_name}', table {table_index}: {discarded_columns_str}")
            
            # If there are non-empty discarded columns (not '', '_1', '_2', etc.), log the dataframe contents
            if non_empty_discarded_columns:
                logger.warning(f"Non-empty discarded columns found: {non_empty_discarded_columns}")
                logger.info(f"DataFrame contents for file '{file_name}', sheet '{sheet_name}', table {table_index}:")
                logger.info("1################################################################################################")
                logger.info(f"{df.to_string()}")
                logger.info("2################################################################################################")
                # Log contents of discarded columns properly
                for col in non_empty_discarded_columns:
                    logger.info(f"Contents of discarded column '{col}':")
                    logger.info(f"{df[col].to_string()}")
                logger.info("3################################################################################################")
        
        if not valid_columns:
            return "Error: No valid columns found in DataFrame that match expected columns"
        
        # Create a filtered dataframe with only valid columns
        df = df[valid_columns]
        
        # Connect to SQLite database using the configured path
        conn = sqlite3.connect(DATABASE_PATH)
        
        # Create table if it doesn't exist with the superset of columns
        create_table_sql = f"""
        CREATE TABLE IF NOT EXISTS payroll_details (
            {', '.join([f'{col} {dtype}' for col, dtype in expected_columns.items()])}
        )
        """
        conn.execute(create_table_sql)
        
        # Prepare the dataframe for insertion
        # Create a copy of the dataframe to avoid SettingWithCopyWarning
        df = df.copy()
        
        # Add missing columns with null values
        for col in expected_columns.keys():
            if col not in df.columns:
                df.loc[:, col] = None
        
        # File name and sheet name are already set by special_logic_preprocess_df
        
        # Ensure correct column order
        df = df[list(expected_columns.keys())]
        
        # Convert data types
        for col, dtype in expected_columns.items():
            if dtype == 'INT':
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            elif dtype == 'NUMERIC(10,2)':
                # Convert to Decimal with 2 decimal places precision
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
                # Round to 2 decimal places using Decimal for precision, then convert to float for SQLite
                df[col] = df[col].apply(lambda x: float(Decimal(str(x)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)))
            else:  # CHAR types
                # For CHAR types, ensure proper string conversion without decimal points for numeric-looking values
                if col == '日期':  # Special handling for date column
                    # Convert to string, but remove '.0' suffix for integer values
                    df[col] = df[col].apply(lambda x: 
                        str(int(float(x))) if pd.notna(x) and str(x).replace('.', '', 1).isdigit() and float(x) == int(float(x)) 
                        else str(x) if pd.notna(x) 
                        else ''
                    )
                else:
                    df[col] = df[col].astype(str)
        
        # Insert data into database (append mode)
        df.to_sql('payroll_details', conn, if_exists='append', index=False)
        
        # Close connection
        conn.close()
        
        # Log the successful operation to log.txt
        success_message = f"Successfully loaded {len(df)} rows to database"
        logger.info(f"File: {file_name}, Sheet: {sheet_name}, Table: {table_index}, Result: {success_message}")
        
        return success_message
        
    except Exception as e:
        # Log the error to log.txt
        error_message = f"Error loading data to database: {str(e)}"
        logger.error(f"File: {file_name}, Sheet: {sheet_name}, Table: {table_index}, Result: {error_message}")
        
        return error_message
